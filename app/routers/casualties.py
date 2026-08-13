"""HTTP router for the Casualties (Втрати) feature.

Exposes:
- GET  /api/cas/units              → ordered list of units
- POST /api/cas/units              → add a new unit
- DELETE /api/cas/units/{id}       → remove a unit (cascades entries)
- POST /api/cas/units/reorder      → update sort order
- GET  /api/cas/entries            → current (undated) working values
- POST /api/cas/entry              → upsert a single entry (no date)
- POST /api/cas/clear-column       → zero one column (no date)
- POST /api/cas/snapshot           → save daily total with date to cas_report_snapshots
- GET  /api/cas/image              → PNG screenshot (date used for header only)
"""

from __future__ import annotations

import re
from datetime import date as _date

from fastapi import APIRouter, Query, Request
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from typing import List, Optional

from app.core.access import current_device_mask
from app.core.db import get_conn
from app.core.time_utils import now_sql

router = APIRouter(tags=["casualties"])


# ── Units ─────────────────────────────────────────────────────────────────────

@router.get("/api/cas/units")
def get_units():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, name, sort_order FROM cas_units ORDER BY sort_order, id"
        ).fetchall()
        return {
            "ok": True,
            "units": [{"id": r["id"], "name": r["name"], "sort_order": r["sort_order"]} for r in rows],
        }


class AddUnitBody(BaseModel):
    name: str


@router.post("/api/cas/units")
def add_unit(body: AddUnitBody):
    name = (body.name or "").strip()
    if not name:
        return JSONResponse({"ok": False, "error": "name required"}, status_code=400)
    with get_conn() as conn:
        max_ord = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) FROM cas_units"
        ).fetchone()[0]
        try:
            row = conn.execute(
                "INSERT INTO cas_units (name, sort_order, created_at) VALUES (?,?,?) RETURNING id",
                (name, int(max_ord) + 1, now_sql()),
            ).fetchone()
            return {"ok": True, "id": row["id"]}
        except Exception as exc:
            return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)


@router.delete("/api/cas/units/{unit_id}")
def delete_unit(unit_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM cas_units WHERE id = ?", (unit_id,))
        return {"ok": True}


class ReorderBody(BaseModel):
    order: List[int]


@router.post("/api/cas/units/reorder")
def reorder_units(body: ReorderBody):
    with get_conn() as conn:
        for idx, uid in enumerate(body.order):
            conn.execute("UPDATE cas_units SET sort_order = ? WHERE id = ?", (idx, uid))
        return {"ok": True}


# ── Entries (undated working values) ──────────────────────────────────────────

@router.get("/api/cas/entries")
def get_entries():
    """Return current working values for all units (no date binding)."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT unit_id, category, morning, night FROM cas_entries"
        ).fetchall()
        entries = {}
        for r in rows:
            entries[f"{r['category']}_{r['unit_id']}"] = {
                "morning": r["morning"],
                "night":   r["night"],
            }
        return {"ok": True, "entries": entries}


class ClearColumnBody(BaseModel):
    column: str   # "morning" | "night"


@router.post("/api/cas/clear-column")
def clear_column(body: ClearColumnBody):
    if body.column not in ("morning", "night"):
        return JSONResponse({"ok": False, "error": "column must be morning or night"}, status_code=400)
    with get_conn() as conn:
        conn.execute(f"UPDATE cas_entries SET {body.column} = 0")
        return {"ok": True}


class SaveEntryBody(BaseModel):
    unit_id:  int
    category: str   # "irr" | "san"
    morning:  int = 0
    night:    int = 0


@router.post("/api/cas/entry")
def save_entry(body: SaveEntryBody):
    if body.category not in ("irr", "san"):
        return JSONResponse({"ok": False, "error": "category must be irr or san"}, status_code=400)
    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO cas_entries (unit_id, category, morning, night)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(unit_id, category)
            DO UPDATE SET morning = excluded.morning, night = excluded.night
            """,
            (body.unit_id, body.category, body.morning, body.night),
        )
        return {"ok": True}


# ── Snapshot (daily total saved with date) ────────────────────────────────────

class SnapshotBody(BaseModel):
    date: str = ""


@router.post("/api/cas/snapshot")
def save_snapshot(body: SnapshotBody):
    """Persist the current working totals (morning+night) to cas_report_snapshots
    under the given date.  Called when the 16-08 report button is pressed.
    Pressing multiple times for the same date overwrites with latest values.
    """
    report_date = (body.date or "").strip() or _date.today().isoformat()
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT e.unit_id, u.name AS unit_name, e.category,
                   COALESCE(e.morning, 0) + COALESCE(e.night, 0) AS total
            FROM cas_entries e
            JOIN cas_units u ON u.id = e.unit_id
            """,
        ).fetchall()
        ts = now_sql()
        for r in rows:
            conn.execute(
                """
                INSERT INTO cas_report_snapshots
                    (report_date, unit_id, unit_name, category, total, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(report_date, unit_id, category)
                DO UPDATE SET
                    unit_name  = excluded.unit_name,
                    total      = excluded.total,
                    created_at = excluded.created_at
                """,
                (report_date, r["unit_id"], r["unit_name"],
                 r["category"], r["total"], ts),
            )
    return {"ok": True, "date": report_date}


# ── Image ─────────────────────────────────────────────────────────────────────

@router.get("/api/cas/image")
def get_cas_image(date: str = Query(default=""), mode: str = Query(default="morning")):
    if mode not in ("morning", "night"):
        return JSONResponse({"ok": False, "error": "mode must be morning or night"}, status_code=400)

    from app.services.cas_image import build_cas_image

    entry_date = (date or "").strip() or _date.today().isoformat()

    with get_conn() as conn:
        unit_rows = conn.execute(
            "SELECT id, name FROM cas_units ORDER BY sort_order, id"
        ).fetchall()
        # Entries are no longer date-keyed — fetch current working values
        entry_rows = conn.execute(
            "SELECT unit_id, category, morning, night FROM cas_entries"
        ).fetchall()

    units = [{"id": r["id"], "name": r["name"]} for r in unit_rows]
    entries_map = {
        f"{r['category']}_{r['unit_id']}": (r["morning"] or 0, r["night"] or 0)
        for r in entry_rows
    }

    buf = build_cas_image(units, entries_map, mode, entry_date)
    filename = f"vtrata-{mode}-{entry_date}.png"
    return StreamingResponse(
        buf,
        media_type="image/png",
        headers={"Content-Disposition": f"inline; filename={filename}"},
    )


# ═══════════════════════════════════════════════════════════════════════════════
# «Втрати 2» — записи втрат, привʼязані до перехоплень (окремо від ручного талію).
#   casualties           — записи-контейнери (статус 200/300, к-сть, причина…)
#   casualty_reasons     — редагований довідник причин
#   casualty_callsigns   — звʼязок запис↔позивний (виставляє life_status позивного)
# ═══════════════════════════════════════════════════════════════════════════════

def _norm_status(s) -> str:
    return "200" if str(s or "").strip() == "200" else "300"


# ── Автопідбір підрозділу (за group_id мережі + ешелоном з networks.unit) ──────
def _flat(s: str) -> str:
    return re.sub(r"\s+", "", (s or "").lower())


def _norm_unit(s: str) -> str:
    s = (s or "").lower().replace("шт3", "штз")   # цифра-3 ≡ кирилиця-З
    s = re.sub(r"25\s*(за|а)\b", "", s)            # 25 ЗА ≡ 25 А
    return re.sub(r"[^а-яіїєґ0-9]", "", s)


def _echelon(unit: str):
    u = (unit or "").lower()
    m = re.search(r"([123])\s*мсб", u)
    if m:
        return m.group(1) + " мсб"
    if re.search(r"шт[3з]", u):
        return "штз"
    if re.search(r"ісб", u):
        return "ісб"
    return None


def _unit_from_net_desc(nd: str) -> str:
    """Юніт із опису р/м перехоплення: 'укх р/м <ЮНІТ> (р-н …)' → '<ЮНІТ>'."""
    m = re.search(r"р/?м\s+(.+?)\s*(?:\(|$)", nd or "", re.IGNORECASE)
    return m.group(1).strip() if m else ""


def _suggest_unit_id(conn, message_id: int):
    """Підбирає cas_unit для перехоплення: юніт беремо з опису р/м САМОГО
    перехоплення (net_description) — він точніший за статичний networks.unit,
    який може відрізнятись від конкретного перехоплення; формування — з group_id."""
    row = conn.execute(
        "SELECT n.unit AS unit, g.name AS grp, m.net_description AS nd FROM messages m "
        "LEFT JOIN networks n ON n.id = m.network_id "
        "LEFT JOIN groups g ON g.id = n.group_id WHERE m.id = ?",
        (message_id,),
    ).fetchone()
    if not row:
        return None, None
    grp = row["grp"] or ""
    unit = _unit_from_net_desc(row["nd"]) or (row["unit"] or "")
    # Немає ні тексту підрозділу, ні відомого формування — підбирати нічого.
    # (Раніше тут був ранній вихід на будь-яку «Невідомо» групу, через що
    # прямий/ббпс-збіг за текстом unit не відпрацьовував — напр. «2 бБпС 71 опБпС».)
    if not unit.strip() and (not grp or grp == "Невідомо"):
        return None, None
    cas = conn.execute("SELECT id, name FROM cas_units").fetchall()
    casn = {_norm_unit(r["name"]): (r["id"], r["name"]) for r in cas}

    def find(name):
        return casn.get(_norm_unit(name))

    known_grp = grp and grp != "Невідомо"
    # 1) Прямий збіг за текстом unit (точний ешелон+формування; ловить і крос-тег,
    #    коли unit явно вказує інше формування, ніж group, і «2 бБпС 71 опБпС»).
    hit = find(unit)
    if hit:
        return hit[0], hit[1]
    # 2) Ешелон із unit + повна назва групи (коли в unit бракує «67 мсд»).
    ech = _echelon(unit)
    if ech and known_grp:
        hit = find(ech + " " + grp)
        if hit:
            return hit[0], hit[1]
    # 3) Варіант (б): інший підрозділ відомого формування → штз формування.
    if known_grp:
        hit = find("штз " + grp)
        if hit:
            return hit[0], hit[1]
    # 4) 2 бБпС за текстом (якщо група «Невідомо», але unit вказує бпс).
    if "ббпс" in _flat(unit):
        for cid, cname in casn.values():
            if "ббпс" in _flat(cname):
                return cid, cname
    # 5) Дивізійний фолбек.
    if re.search(r"67\s*мсд", (unit + " " + grp).lower()):
        hit = find("67 мсд")
        if hit:
            return hit[0], hit[1]
    return None, None


# ── Довідник причин ────────────────────────────────────────────────────────────

@router.get("/api/casualty-reasons")
def cas_reasons_list():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, name FROM casualty_reasons WHERE is_active = 1 ORDER BY sort_order, id"
        ).fetchall()
    return {"ok": True, "reasons": [{"id": r["id"], "name": r["name"]} for r in rows]}


class ReasonBody(BaseModel):
    name: str


@router.post("/api/casualty-reasons")
def cas_reason_add(body: ReasonBody):
    name = (body.name or "").strip()
    if not name:
        return JSONResponse({"ok": False, "error": "name required"}, status_code=400)
    with get_conn() as conn:
        existing = conn.execute(
            "SELECT id FROM casualty_reasons WHERE name = ?", (name,)
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE casualty_reasons SET is_active = 1 WHERE id = ?", (existing["id"],)
            )
            return {"ok": True, "id": existing["id"]}
        mx = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) FROM casualty_reasons"
        ).fetchone()[0]
        row = conn.execute(
            "INSERT INTO casualty_reasons (name, sort_order) VALUES (?, ?) RETURNING id",
            (name, int(mx) + 1),
        ).fetchone()
    return {"ok": True, "id": row["id"]}


# ── Записи втрат ───────────────────────────────────────────────────────────────

def _casualty_dict(conn, r) -> dict:
    """Серіалізує рядок casualties (з join reason_name/unit_name) + позивні."""
    cs = conn.execute(
        "SELECT c.id, c.name FROM casualty_callsigns cc "
        "JOIN callsigns c ON c.id = cc.callsign_id WHERE cc.casualty_id = ?",
        (r["id"],),
    ).fetchall()
    keys = r.keys()
    unit_name = r["unit_name"] if "unit_name" in keys else None
    return {
        "id": r["id"],
        "message_id": r["message_id"],
        "status": r["status"],
        "count": r["count"],
        "reason_id": r["reason_id"],
        "reason": (r["reason_name"] if "reason_name" in keys else None),
        "unit_id": r["unit_id"],
        "unit": (unit_name or r["unit_text"] or ""),
        "unit_text": r["unit_text"],
        "accounted": bool(r["accounted"]),
        "comment": r["comment"] or "",
        "created_at": r["created_at"],
        "created_by": r["created_by"],
        "last_edited_by": r["last_edited_by"],
        "last_edited_at": r["last_edited_at"],
        "callsigns": [{"id": x["id"], "name": x["name"]} for x in cs],
        # Контекст перехоплення (для сторінкового переліку «Втрати 2»).
        "msg_dt": (r["msg_dt"] if "msg_dt" in keys else None),
        "msg_freq": (r["msg_freq"] if "msg_freq" in keys else None),
        "msg_net": (r["msg_net"] if "msg_net" in keys else None),
        "msg_text": (r["msg_text"] if "msg_text" in keys else None),
        "msg_network_id": (r["msg_network_id"] if "msg_network_id" in keys else None),
    }


_SELECT_CASUALTY = (
    "SELECT cas.*, r.name AS reason_name, u.name AS unit_name, "
    "m.created_at AS msg_dt, m.net_description AS msg_net, m.body_text AS msg_text, "
    "n.frequency AS msg_freq, n.id AS msg_network_id "
    "FROM casualties cas "
    "LEFT JOIN casualty_reasons r ON r.id = cas.reason_id "
    "LEFT JOIN cas_units u        ON u.id = cas.unit_id "
    "LEFT JOIN messages m         ON m.id = cas.message_id "
    "LEFT JOIN networks n         ON n.id = m.network_id "
)


def _fmt_intercept_dt(s: str) -> str:
    v = str(s or "").replace("T", " ").strip()
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})[ ](\d{2}:\d{2}:\d{2})", v)
    return f"{m.group(3)}.{m.group(2)}.{m.group(1)}, {m.group(4)}" if m else v


def build_standard_intercept_text(conn, message_id: int) -> str:
    """Перехоплення за стандартним шаблоном (як кнопка «Копіювати»):
    дата/час, маска|частота, опис мережі, позивні-ініціатори (caller),
    позивні-адресати (callee/mentioned), порожній рядок, текст діалогу."""
    row = conn.execute(
        "SELECT m.created_at AS dt, m.net_description AS net, m.body_text AS body, "
        "n.frequency AS freq, n.mask AS mask "
        "FROM messages m LEFT JOIN networks n ON n.id = m.network_id WHERE m.id = ?",
        (message_id,),
    ).fetchone()
    if not row:
        return ""
    callers: list = []
    callees: list = []
    for r in conn.execute(
        "SELECT mc.role AS role, c.name AS name FROM message_callsigns mc "
        "JOIN callsigns c ON c.id = mc.callsign_id WHERE mc.message_id = ? ORDER BY c.name",
        (message_id,),
    ).fetchall():
        nm = (r["name"] or "").strip()
        if not nm:
            continue
        if r["role"] == "caller":
            callers.append(nm)
        elif r["role"] in ("callee", "mentioned"):
            callees.append(nm)
    mask = (row["mask"] or "").strip()
    freq = (row["freq"] or "").strip()
    parts = [
        _fmt_intercept_dt(row["dt"]),
        mask if mask else freq,      # 2-й рядок — маска, якщо є, інакше частота
        (row["net"] or "").strip(),
        ", ".join(callers),
        ", ".join(callees),
    ]
    body = (row["body"] or "").rstrip()
    if body:
        parts.append("")
        parts.append(body)
    return "\n".join(parts)


def _apply_callsign_casualty(conn, callsign_id: int, status: str,
                             message_id: int, editor: str, ts: str) -> None:
    """Прив'язаному позивному виставляє life_status (200/300) і дописує
    ПОВНЕ перехоплення за стандартним шаблоном у коментар (без дублювання).
    Якщо в коментарі вже є текст — відділяємо новий блок порожнім рядком."""
    row = conn.execute("SELECT comment FROM callsigns WHERE id = ?", (callsign_id,)).fetchone()
    if not row:
        return
    comment = row["comment"] or ""
    intercept_text = build_standard_intercept_text(conn, message_id).strip()
    if intercept_text and intercept_text not in comment:
        comment = comment.rstrip() + ("\n\n" if comment.strip() else "") + intercept_text
    conn.execute(
        "UPDATE callsigns SET life_status = ?, comment = ?, "
        "last_edited_by = ?, last_edited_at = ? WHERE id = ?",
        (status, comment, editor, ts, callsign_id),
    )


def _recompute_callsign_status(conn, callsign_id: int, editor: str, ts: str) -> None:
    """Перераховує life_status позивного за НАЯВНИМИ валідними записами втрат.
    200 має пріоритет над 300; якщо записів не лишилось — повертаємо 'alive'.
    Викликається після видалення запису втрат, щоб зняти помилковий 200/300."""
    row = conn.execute(
        "SELECT cas.status AS status FROM casualty_callsigns cc "
        "JOIN casualties cas ON cas.id = cc.casualty_id "
        "WHERE cc.callsign_id = ? AND cas.is_valid = 1 "
        "ORDER BY CASE cas.status WHEN '200' THEN 0 ELSE 1 END LIMIT 1",
        (callsign_id,),
    ).fetchone()
    new_status = row["status"] if row else "alive"
    conn.execute(
        "UPDATE callsigns SET life_status = ?, last_edited_by = ?, last_edited_at = ? WHERE id = ?",
        (new_status, editor, ts, callsign_id),
    )


@router.get("/api/casualties")
def cas_records_list(
    message_id: int = Query(default=0),
    date_from: str = Query(default=""),
    date_to: str = Query(default=""),
):
    """Записи втрат: для конкретного перехоплення (message_id) або за періодом."""
    wheres = ["cas.is_valid = 1"]
    params: list = []
    if message_id:
        wheres.append("cas.message_id = ?"); params.append(message_id)
    # Період — за часом ПЕРЕХОПЛЕННЯ (messages.created_at), не за часом внесення.
    # Нормалізуємо роздільник дати: інпут дає 'T' (2026-08-02T16:00), а
    # created_at зберігається з пробілом — а ' ' (32) < 'T' (84), тож без
    # нормалізації записи того ж дня, що й нижня межа, помилково відкидались.
    if date_from:
        wheres.append("REPLACE(m.created_at,'T',' ') >= ?"); params.append(date_from.replace("T", " "))
    if date_to:
        wheres.append("REPLACE(m.created_at,'T',' ') <= ?"); params.append(date_to.replace("T", " "))
    sql = _SELECT_CASUALTY + "WHERE " + " AND ".join(wheres) + \
        " ORDER BY m.created_at DESC, cas.id DESC"
    with get_conn() as conn:
        rows = conn.execute(sql, params).fetchall()
        out = [_casualty_dict(conn, r) for r in rows]
        resp = {"ok": True, "records": out}
        if message_id:
            uid, uname = _suggest_unit_id(conn, message_id)
            resp["suggested_unit_id"] = uid
            resp["suggested_unit_name"] = uname
            resp["standard_text"] = build_standard_intercept_text(conn, message_id)
    return resp


def _aggregate_report(date_from: str, date_to: str):
    """Агрегація втрат за період по підрозділах × статус (200/300).
    Враховані (accounted=1) — не входять. Повертає (rows, totals)."""
    wheres = ["cas.is_valid = 1", "COALESCE(cas.accounted, 0) = 0"]
    params: list = []
    # Нормалізуємо роздільник дати: інпут дає 'T' (2026-08-02T16:00), а
    # created_at зберігається з пробілом — а ' ' (32) < 'T' (84), тож без
    # нормалізації записи того ж дня, що й нижня межа, помилково відкидались.
    if date_from:
        wheres.append("REPLACE(m.created_at,'T',' ') >= ?"); params.append(date_from.replace("T", " "))
    if date_to:
        wheres.append("REPLACE(m.created_at,'T',' ') <= ?"); params.append(date_to.replace("T", " "))
    sql = (
        "SELECT COALESCE(NULLIF(u.name, ''), NULLIF(cas.unit_text, ''), 'Невідомо') AS unit, "
        "u.sort_order AS so, cas.status AS status, SUM(cas.count) AS cnt "
        "FROM casualties cas LEFT JOIN cas_units u ON u.id = cas.unit_id "
        "LEFT JOIN messages m ON m.id = cas.message_id "
        "WHERE " + " AND ".join(wheres) + " GROUP BY unit, cas.status"
    )
    reason_sql = (
        "SELECT COALESCE(NULLIF(r.name, ''), 'Невідомо') AS reason, SUM(cas.count) AS cnt "
        "FROM casualties cas LEFT JOIN casualty_reasons r ON r.id = cas.reason_id "
        "LEFT JOIN messages m ON m.id = cas.message_id "
        "WHERE " + " AND ".join(wheres) + " GROUP BY reason ORDER BY cnt DESC"
    )
    agg: dict = {}
    by_reason: list = []
    with get_conn() as conn:
        for r in conn.execute(sql, params).fetchall():
            u = r["unit"]
            row = agg.setdefault(u, {
                "unit": u, "killed": 0, "wounded": 0,
                "so": r["so"] if r["so"] is not None else 9999,
            })
            if r["status"] == "200":
                row["killed"] += r["cnt"] or 0
            else:
                row["wounded"] += r["cnt"] or 0
        by_reason = [
            {"reason": rr["reason"], "cnt": rr["cnt"] or 0}
            for rr in conn.execute(reason_sql, params).fetchall() if (rr["cnt"] or 0) > 0
        ]
    rows = sorted(agg.values(), key=lambda x: (x["so"], x["unit"]))
    totals = {
        "killed": sum(x["killed"] for x in rows),
        "wounded": sum(x["wounded"] for x in rows),
    }
    return rows, totals, by_reason


@router.get("/api/casualties/report")
def cas_report(date_from: str = Query(default=""), date_to: str = Query(default="")):
    """Агрегація втрат за період: по підрозділах × статус (200/300) + за причинами."""
    rows, totals, by_reason = _aggregate_report(date_from, date_to)
    return {"ok": True, "rows": rows, "totals": totals, "by_reason": by_reason}


def _fmt_period_label(date_from: str, date_to: str) -> str:
    def one(s):
        v = str(s or "").replace("T", " ").strip()
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})[ ](\d{2}:\d{2})", v)
        return f"{m.group(3)}.{m.group(2)}.{m.group(1)} {m.group(4)}" if m else v
    a, b = one(date_from), one(date_to)
    if a and b:
        return f"{a} – {b}"
    return a or b or ""


@router.get("/api/casualties/report-image")
def cas_report_image(
    date_from: str = Query(default=""), date_to: str = Query(default=""),
    date_from2: str = Query(default=""), date_to2: str = Query(default=""),
    col1: str = Query(default=""), col2: str = Query(default=""),
):
    """PNG-зображення звіту втрат за період (кнопка «фотоапарат» у модалці).
    Якщо задано другий період (date_from2/date_to2) — звіт із ДВОМА колонками."""
    from app.services.cas_image import build_casualty_report_image, build_casualty_report_image_2col

    if date_from2 or date_to2:
        rows1, tot1, _ = _aggregate_report(date_from, date_to)
        rows2, tot2, _ = _aggregate_report(date_from2, date_to2)

        def _merge(key):
            m: dict = {}
            for r in rows1:
                if r[key]:
                    m.setdefault(r["unit"], {"unit": r["unit"], "so": r["so"], "c1": 0, "c2": 0})["c1"] += r[key]
            for r in rows2:
                if r[key]:
                    m.setdefault(r["unit"], {"unit": r["unit"], "so": r["so"], "c1": 0, "c2": 0})["c2"] += r[key]
            return [(x["unit"], x["c1"], x["c2"]) for x in sorted(m.values(), key=lambda z: (z["so"], z["unit"]))]

        sections = [
            ("irr", "БЕЗПОВОРОТНІ ВТРАТИ", _merge("killed")),
            ("san", "САНІТАРНІ ВТРАТИ", _merge("wounded")),
        ]
        buf2 = build_casualty_report_image_2col(
            sections, _fmt_period_label(date_from2 or date_from, date_to2 or date_to),
            col1 or "16:00–08:00", col2 or "08:00–08:00", tot1, tot2,
        )
        return StreamingResponse(buf2, media_type="image/png",
                                 headers={"Content-Disposition": "inline; filename=vtraty-2col.png"})

    rows, totals, _ = _aggregate_report(date_from, date_to)
    irr_items = [(x["unit"], x["killed"]) for x in rows if x["killed"]]
    san_items = [(x["unit"], x["wounded"]) for x in rows if x["wounded"]]
    sections = [
        ("irr", "БЕЗПОВОРОТНІ ВТРАТИ", irr_items),
        ("san", "САНІТАРНІ ВТРАТИ", san_items),
    ]
    buf = build_casualty_report_image(
        sections, _fmt_period_label(date_from, date_to),
        totals["killed"], totals["wounded"],
    )
    return StreamingResponse(
        buf, media_type="image/png",
        headers={"Content-Disposition": "inline; filename=vtraty-2.png"},
    )


@router.get("/api/casualties/report.xlsx")
def cas_report_xlsx(date_from: str = Query(default=""), date_to: str = Query(default="")):
    """XLSX звіту втрат за період: таблиця (Підрозділ | 200 | 300 | Усього),
    а під нею дві кругові діаграми — розподіл за підрозділами та за причинами.
    Кнопка «фотоапарат» у модалці з затиснутим Ctrl."""
    from io import BytesIO
    import openpyxl
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    rows, totals, by_reason = _aggregate_report(date_from, date_to)
    period = _fmt_period_label(date_from, date_to)

    FNAME = "Times New Roman"
    f_norm  = Font(name=FNAME, size=14)
    f_bold  = Font(name=FNAME, size=14, bold=True)
    f_hdr   = Font(name=FNAME, size=14, bold=True, color="FFFFFF")
    f_title = Font(name=FNAME, size=15, bold=True)
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    hdr_fill = PatternFill("solid", fgColor="1F2937")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Втрати"

    ws["A1"] = f"Втрати за період: {period}" if period else "Втрати за період"
    ws["A1"].font = f_title
    ws.merge_cells("A1:B1")

    def _hcell(row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = f_hdr; c.fill = hdr_fill; c.border = box; c.alignment = center
        return c

    HDR = 3
    # Основна таблиця — 2 колонки: Підрозділ | Кількість (сумарні втрати підрозділу).
    _hcell(HDR, 1, "Підрозділ")
    _hcell(HDR, 2, "Кількість")
    r = HDR + 1
    unit_first = r
    for row in rows:
        tot = int(row["killed"] or 0) + int(row["wounded"] or 0)
        a = ws.cell(row=r, column=1, value=row["unit"]); a.font = f_norm; a.border = box; a.alignment = left
        b = ws.cell(row=r, column=2, value=tot); b.font = f_norm; b.border = box; b.alignment = center
        r += 1
    unit_last = r - 1
    ua = ws.cell(row=r, column=1, value="Усього"); ua.font = f_bold; ua.border = box; ua.alignment = left
    ub = ws.cell(row=r, column=2, value=int(totals["killed"]) + int(totals["wounded"]))
    ub.font = f_bold; ub.border = box; ub.alignment = center
    total_row = r

    # Дані для діаграми «за причинами» — допоміжна табличка праворуч (D/E).
    _hcell(HDR, 4, "Причина")
    _hcell(HDR, 5, "Кількість")
    rr = HDR + 1
    reason_first = rr
    for br in by_reason:
        d = ws.cell(row=rr, column=4, value=br["reason"]); d.font = f_norm; d.border = box; d.alignment = left
        e = ws.cell(row=rr, column=5, value=int(br["cnt"] or 0)); e.font = f_norm; e.border = box; e.alignment = center
        rr += 1
    reason_last = rr - 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 13

    # Читабельні діаграми: на секторах — ЛИШЕ відсоток (короткий), а назви
    # підрозділів/причин — у легенді праворуч (довгі підписи не налазять один
    # на одного, як було з showCatName + showSerName).
    def _pie(title, val_col, cat_col, first, last, anchor):
        p = PieChart()
        p.title = title
        p.add_data(Reference(ws, min_col=val_col, min_row=first, max_row=last), titles_from_data=False)
        p.set_categories(Reference(ws, min_col=cat_col, min_row=first, max_row=last))
        dl = DataLabelList()
        dl.showSerName = False
        dl.showCatName = False
        dl.showVal = False
        dl.showPercent = True
        dl.showLegendKey = False
        dl.numFmt = "0%"
        p.dataLabels = dl
        if p.legend is not None:
            p.legend.position = "r"
            p.legend.overlay = False
        p.height = 10.5
        p.width = 20
        ws.add_chart(p, anchor)

    anchor_row = total_row + 3
    if unit_last >= unit_first:
        _pie("Розподіл втрат за підрозділами", 2, 1, unit_first, unit_last, f"A{anchor_row}")
    if reason_last >= reason_first:
        _pie("Розподіл втрат за причинами", 5, 4, reason_first, reason_last, f"A{anchor_row + 23}")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="vtraty.xlsx"'},
    )


class CasualtyIn(BaseModel):
    message_id: int
    status: str = "300"
    count: int = 1
    reason_id: Optional[int] = None
    unit_id: Optional[int] = None
    unit_text: str = ""
    accounted: bool = False
    comment: str = ""
    callsign_ids: List[int] = []
    # Чи виставляти life_status позивним. Інкрементальне збереження — False
    # (лише фіксуємо звʼязок); застосування статусів — на кнопці «Надіслати».
    apply_status: bool = True


@router.post("/api/casualties")
def cas_record_create(body: CasualtyIn, request: Request):
    if not body.message_id:
        return JSONResponse({"ok": False, "error": "message_id required"}, status_code=400)
    status = _norm_status(body.status)
    editor = current_device_mask(request)
    ts = now_sql()
    with get_conn() as conn:
        new = conn.execute(
            """
            INSERT INTO casualties
                (message_id, status, count, reason_id, unit_id, unit_text, accounted,
                 comment, created_at, created_by, last_edited_at, last_edited_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            RETURNING id
            """,
            (body.message_id, status, max(1, int(body.count or 1)), body.reason_id,
             body.unit_id, (body.unit_text or "").strip(), 1 if body.accounted else 0,
             (body.comment or "").strip(), ts, editor, ts, editor),
        ).fetchone()
        cas_id = new["id"]
        for cid in dict.fromkeys(body.callsign_ids or []):
            conn.execute(
                "INSERT OR IGNORE INTO casualty_callsigns (casualty_id, callsign_id) VALUES (?, ?)",
                (cas_id, cid),
            )
            if body.apply_status:
                _apply_callsign_casualty(conn, cid, status, body.message_id, editor, ts)
        r = conn.execute(_SELECT_CASUALTY + "WHERE cas.id = ?", (cas_id,)).fetchone()
        out = _casualty_dict(conn, r)
    return {"ok": True, "record": out}


class CasualtyUpdate(BaseModel):
    status: Optional[str] = None
    count: Optional[int] = None
    reason_id: Optional[int] = None
    unit_id: Optional[int] = None
    unit_text: Optional[str] = None
    accounted: Optional[bool] = None
    comment: Optional[str] = None
    callsign_ids: Optional[List[int]] = None
    apply_status: bool = False   # застосовувати life_status лише на «Надіслати»


@router.put("/api/casualties/{cas_id}")
def cas_record_update(cas_id: int, body: CasualtyUpdate, request: Request):
    editor = current_device_mask(request)
    ts = now_sql()
    with get_conn() as conn:
        cur = conn.execute(
            "SELECT * FROM casualties WHERE id = ? AND is_valid = 1", (cas_id,)
        ).fetchone()
        if not cur:
            return JSONResponse({"ok": False, "error": "not found"}, status_code=404)
        status    = _norm_status(body.status) if body.status is not None else cur["status"]
        count     = max(1, int(body.count)) if body.count is not None else cur["count"]
        reason_id = body.reason_id if body.reason_id is not None else cur["reason_id"]
        unit_id   = body.unit_id if body.unit_id is not None else cur["unit_id"]
        unit_text = (body.unit_text or "").strip() if body.unit_text is not None else cur["unit_text"]
        accounted = (1 if body.accounted else 0) if body.accounted is not None else cur["accounted"]
        comment   = (body.comment or "").strip() if body.comment is not None else cur["comment"]
        conn.execute(
            "UPDATE casualties SET status = ?, count = ?, reason_id = ?, unit_id = ?, "
            "unit_text = ?, accounted = ?, comment = ?, last_edited_at = ?, last_edited_by = ? "
            "WHERE id = ?",
            (status, count, reason_id, unit_id, unit_text, accounted, comment, ts, editor, cas_id),
        )
        removed_ids: List[int] = []
        if body.callsign_ids is not None:
            # Запам'ятовуємо стару прив'язку, щоб знайти ПРИБРАНІ позивні —
            # їм треба перерахувати статус (зняти помилковий 200/300).
            old_ids = [r["callsign_id"] for r in conn.execute(
                "SELECT callsign_id FROM casualty_callsigns WHERE casualty_id = ?", (cas_id,)
            ).fetchall()]
            new_ids = list(dict.fromkeys(body.callsign_ids))
            conn.execute("DELETE FROM casualty_callsigns WHERE casualty_id = ?", (cas_id,))
            for cid in new_ids:
                conn.execute(
                    "INSERT OR IGNORE INTO casualty_callsigns (casualty_id, callsign_id) VALUES (?, ?)",
                    (cas_id, cid),
                )
            new_set = set(new_ids)
            removed_ids = [c for c in old_ids if c not in new_set]
        # Прибрані позивні перераховуємо ЗАВЖДИ — навіть при інкрементальному
        # збереженні (apply_status=False). Відв'язаний позивний більше не втрата,
        # тож знімаємо помилковий 200/300 (немає інших валідних записів → 'alive').
        # Без цього explicit-save не бачив би зміни: інкрементальний автозбереж
        # уже відв'язав позивного, і removed_ids на «Зберегти» був би порожнім.
        for rid in removed_ids:
            _recompute_callsign_status(conn, rid, editor, ts)
        # Виставлення 200/300 привʼязаним — лише коли просять (Зберегти/Надіслати).
        if body.apply_status:
            linked = conn.execute(
                "SELECT callsign_id FROM casualty_callsigns WHERE casualty_id = ?", (cas_id,)
            ).fetchall()
            for lr in linked:
                _apply_callsign_casualty(conn, lr["callsign_id"], status, cur["message_id"], editor, ts)
        r = conn.execute(_SELECT_CASUALTY + "WHERE cas.id = ?", (cas_id,)).fetchone()
        out = _casualty_dict(conn, r)
    return {"ok": True, "record": out}


@router.delete("/api/casualties/{cas_id}")
def cas_record_delete(cas_id: int, request: Request):
    """Мʼяке видалення (is_valid=0). Після видалення перераховуємо life_status
    привʼязаних позивних: якщо інших валідних записів втрат немає — знімаємо
    200/300 (повертаємо 'alive'); коментар при цьому не чіпаємо."""
    editor = current_device_mask(request)
    ts = now_sql()
    with get_conn() as conn:
        linked = conn.execute(
            "SELECT callsign_id FROM casualty_callsigns WHERE casualty_id = ?", (cas_id,)
        ).fetchall()
        conn.execute(
            "UPDATE casualties SET is_valid = 0, last_edited_at = ?, last_edited_by = ? WHERE id = ?",
            (ts, editor, cas_id),
        )
        for lr in linked:
            _recompute_callsign_status(conn, lr["callsign_id"], editor, ts)
    return {"ok": True}
