"""UI and API router for managing radio networks.

This router serves:

- HTML pages for viewing/searching/editing networks (`/networks`);
- small JSON API helpers used by frontend code for lookups.

The router is primarily a presentation layer: it loads reference data
(statuses/chats/groups/tags), renders templates, and executes simple SQL
updates. Complex domain rules such as network resolution during ingest are
handled in service modules (see `app.services.network_service`).
"""

from __future__ import annotations

import logging
import sqlite3
import threading
from datetime import datetime, date, timedelta
from typing import List, Optional

log = logging.getLogger(__name__)

from app.services.network_search import search_network_rows

from fastapi import APIRouter, Depends, Request, Form
from fastapi.responses import RedirectResponse, HTMLResponse, JSONResponse, Response
from pydantic import BaseModel

from app.core.config import settings
from app.core.db import get_conn
from app.core.normalize import normalize_freq, normalize_freq_or_mask
from app.core.auth_context import get_actor

router = APIRouter()


class NetworkAliasCreateIn(BaseModel):
    network_id: int
    alias_text: str


class NetworkAliasUpdateIn(BaseModel):
    alias_text: str
    network_id: Optional[int] = None


class NetworkVerifyIn(BaseModel):
    items: list[str]


def _clean_alias_text(raw: str) -> str:
    """Trim trailing whitespace from user-entered alias (preserve leading for rare cases)."""
    return (raw or "").rstrip()


def _status_has_colors(conn) -> bool:
    """Return True if `statuses` table supports bg/border color columns."""
    cols = conn.execute("PRAGMA table_info(statuses)").fetchall()
    col_names = [c[1] for c in cols]
    return "bg_color" in col_names and "border_color" in col_names


def _all_networks_list(conn, status_ids, chat_ids, group_ids):
    """List networks for the "Усі р/м" tab with optional filters."""
    has_colors = _status_has_colors(conn)
    select_colors = (
        "s.bg_color, s.border_color"
        if has_colors
        else "NULL as bg_color, NULL as border_color"
    )

    base_sql = f"""
    SELECT
        n.id,
        n.frequency,
        n.mask,
        n.unit,
        n.zone,
        c.name AS chat_name,
        g.name AS group_name,
        s.name AS status_name,
        {select_colors},
        e.updated_at AS etalon_updated_at,
        (SELECT COUNT(*) FROM messages m
           WHERE m.network_id = n.id
             AND m.is_valid = 1
             AND coalesce(m.content_type, 'intercept') = 'intercept'
             AND REPLACE(m.created_at, 'T', ' ') >= ?) AS intercepts_7d
    FROM networks n
    JOIN chats c    ON c.id = n.chat_id
    JOIN groups g   ON g.id = n.group_id
    JOIN statuses s ON s.id = n.status_id
    LEFT JOIN etalons e ON e.network_id = n.id
    """

    clauses = []
    # First bind param feeds the intercepts_7d subquery in the SELECT above.
    since_7d = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
    params = [since_7d]

    def add_in(field, values):
        if not values:
            return
        values = [int(v) for v in values if int(v) != 0]
        if not values:
            return
        placeholders = ",".join(["?"] * len(values))
        clauses.append(f"{field} IN ({placeholders})")
        params.extend(values)

    add_in("n.status_id", status_ids)
    add_in("n.chat_id", chat_ids)
    add_in("n.group_id", group_ids)

    if clauses:
        base_sql += " WHERE " + " AND ".join(clauses)

    base_sql += " ORDER BY intercepts_7d DESC, n.frequency ASC"
    rows = conn.execute(base_sql, params).fetchall()

    # Прикріпити теги (як у таблиці на головній): network_id -> [tag_id, ...]
    tag_map = _get_tag_ids_for_networks(conn, [int(r["id"]) for r in rows])
    out = []
    for r in rows:
        d = dict(r)
        d["tag_ids"] = tag_map.get(int(r["id"]), [])
        out.append(d)
    return out


def _default_all_status_ids(conn) -> list[int]:
    rows = conn.execute(
        """
        SELECT id
        FROM statuses
        WHERE name = 'Спостерігається'
        """
    ).fetchall()
    return [int(r[0]) for r in rows] if rows else [0]


@router.get("/api/networks/{network_id}/callsign-graph")
def api_network_callsign_graph(network_id: int, days: int = 14):
    """Return callsign graph (nodes+edges) for a network within last N days."""
    try:
        nid = int(network_id)
    except Exception:
        return {"ok": False, "error": "invalid network_id"}

    try:
        days_n = int(days)
        if days_n < 1:
            days_n = 1
        if days_n > 365:
            days_n = 365
    except Exception:
        days_n = 14

    start_dt = (datetime.now() - timedelta(days=days_n)).isoformat(timespec="seconds")

    with get_conn() as conn:
        # Nodes: callsigns seen in the period.
        node_rows = conn.execute(
            """
            SELECT
              c.id,
              c.name,
              COALESCE(c.callsign_status_id, c.status_id) AS status_id
            FROM callsigns c
            WHERE c.network_id = ?
              AND COALESCE(c.last_seen_dt, '') >= ?
              AND c.name <> 'НВ'
            ORDER BY c.name
            """,
            (nid, start_dt),
        ).fetchall()

        node_ids = [int(r["id"]) for r in node_rows]
        if not node_ids:
            return {"ok": True, "nodes": [], "edges": [], "meta": {"days": days_n}}

        placeholders = ",".join(["?"] * len(node_ids))

        edge_rows = conn.execute(
            f"""
            SELECT
              a_callsign_id AS a_id,
              b_callsign_id AS b_id,
              cnt,
              last_seen_dt
            FROM callsign_edges
            WHERE network_id = ?
              AND last_seen_dt >= ?
              AND a_callsign_id IN ({placeholders})
              AND b_callsign_id IN ({placeholders})
            ORDER BY cnt DESC, last_seen_dt DESC
            """,
            (nid, start_dt, *node_ids, *node_ids),
        ).fetchall()

    nodes = []
    for r in node_rows:
        sid = r["status_id"]
        sid_num = int(sid) if sid is not None and str(sid).strip() != "" else None
        icon = f"/static/icons/callsign_statuses/{sid_num}.svg" if sid_num else "/static/icons/callsign_statuses/_default.svg"
        nodes.append(
            {
                "id": int(r["id"]),
                "name": r["name"] or "",
                "status_id": sid_num,
                "icon": icon,
            }
        )

    edges = []
    for r in edge_rows:
        edges.append(
            {
                "source": int(r["a_id"]),
                "target": int(r["b_id"]),
                "cnt": int(r["cnt"] or 0),
            }
        )

    return {"ok": True, "nodes": nodes, "edges": edges, "meta": {"days": days_n, "start_dt": start_dt}}


@router.get("/api/networks/{network_id}/callsigns.xlsx")
def api_network_callsigns_xlsx(network_id: int):
    """Експорт позивних радіомережі у .xlsx.

    Колонки: н/п, позивний, іконка (PNG статусу), роль (caller/callee), коментар.
    У заголовку таблиці та назві файлу — частота, підрозділ, зона функціонування.
    """
    import io
    import re as _re
    from pathlib import Path as _Path
    from urllib.parse import quote as _quote

    import cairosvg
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.units import pixels_to_EMU

    with get_conn() as conn:
        net = conn.execute(
            "SELECT id, frequency, mask, unit, zone FROM networks WHERE id=?",
            (network_id,),
        ).fetchone()
        if not net:
            return JSONResponse({"ok": False, "error": "Радіомережу не знайдено"}, status_code=404)

        rows = conn.execute(
            "SELECT c.id, c.name, c.comment, c.callsign_status_id AS sid, "
            "       cs.name AS role_name "
            "FROM callsigns c "
            "LEFT JOIN callsign_statuses cs ON cs.id = c.callsign_status_id "
            "WHERE c.network_id=? AND c.name <> 'НВ' "
            "ORDER BY c.name COLLATE NOCASE",
            (network_id,),
        ).fetchall()

    freq = (net["frequency"] or "").strip()
    unit = (net["unit"] or "").strip()
    zone = (net["zone"] or "").strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Позивні"

    # Іконка = SVG «Швидкої ідентифікації» статусу позивного
    # (app/static/icons/callsign_statuses/{callsign_status_id}.svg).
    # xlsx вимагає растр, тож рендеримо SVG → PNG 32×32 (cairosvg) один раз
    # на кожен унікальний статус і кешуємо по status_id.
    icons_dir = _Path(__file__).resolve().parent.parent / "static" / "icons" / "callsign_statuses"
    _icon_cache: dict = {}

    def _icon_png(sid) -> bytes | None:
        if sid is None or str(sid).strip() == "":
            return None
        key = int(sid)
        if key in _icon_cache:
            return _icon_cache[key]
        path = icons_dir / f"{key}.svg"
        data = None
        if path.exists():
            try:
                data = cairosvg.svg2png(url=str(path), output_width=32, output_height=32)
            except Exception:
                data = None
        _icon_cache[key] = data
        return data

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="DDDDDD")

    # Заголовок-рядок із частотою / підрозділом / зоною.
    title = "Позивні р/м — " + " | ".join(p for p in (freq, unit, zone) if p)
    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    headers = ["н/п", "позивний", "іконка", "роль", "коментар"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    widths = {"A": 6, "B": 24, "C": 8, "D": 18, "E": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    rownum = 3
    for i, row in enumerate(rows, start=1):
        ws.cell(row=rownum, column=1, value=i).alignment = Alignment(horizontal="center", vertical="center")
        c_name = ws.cell(row=rownum, column=2, value=row["name"] or "")
        c_name.alignment = Alignment(horizontal="center", vertical="center")
        c_name.font = Font(bold=True)
        ws.cell(row=rownum, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=rownum, column=4, value=row["role_name"] or "").alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=rownum, column=5, value=row["comment"] or "").alignment = Alignment(vertical="center", wrap_text=True)
        for col in range(1, 6):
            ws.cell(row=rownum, column=col).border = border
        ws.row_dimensions[rownum].height = 26

        png_bytes = _icon_png(row["sid"])
        if png_bytes:
            try:
                img = XLImage(io.BytesIO(png_bytes))
                # Центруємо іконку 32×32 у клітинці C (плаваюче зображення не
                # реагує на вирівнювання клітинки — рахуємо відступи вручну).
                icon_px = 32
                col_w_px = int(widths["C"] * 7 + 5)   # ширина стовпця у px
                row_h_px = int(26 * 96 / 72)           # висота рядка (26pt) у px
                col_off = max(0, (col_w_px - icon_px) // 2)
                row_off = max(0, (row_h_px - icon_px) // 2)
                marker = AnchorMarker(
                    col=2, colOff=pixels_to_EMU(col_off),
                    row=rownum - 1, rowOff=pixels_to_EMU(row_off),
                )
                img.anchor = OneCellAnchor(
                    _from=marker,
                    ext=XDRPositiveSize2D(pixels_to_EMU(icon_px), pixels_to_EMU(icon_px)),
                )
                ws.add_image(img)
            except Exception:
                pass
        rownum += 1

    bio = io.BytesIO()
    wb.save(bio)

    raw = " ".join(p for p in ("Позивні", freq, unit, zone) if p)
    safe = _re.sub(r"\s+", " ", _re.sub(r'[\\/:*?"<>|]+', " ", raw)).strip() or "Позивні"
    fn = safe + ".xlsx"
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''" + _quote(fn)},
    )


@router.get("/api/networks/{network_id}/intercept-stats")
def api_network_intercept_stats(network_id: int, days: int = 7):
    """Daily intercept counts for a network over the last `days` days.

    "Intercept" uses the same definition as the home page: messages with
    is_valid = 1 and content_type = 'intercept' (analytical/peleng excluded),
    bucketed by the intercept datetime (`created_at`) date.

    Returns per-date counts only (no zero-filling). `created_at` is stored in
    local time, so the client builds the fixed N-day axis in the user's local
    timezone and maps these counts onto it; a couple of extra days are included
    here as a buffer for timezone edge effects.
    """
    try:
        nid = int(network_id)
    except Exception:
        return {"ok": False, "error": "invalid network_id"}

    try:
        days_n = int(days)
        days_n = max(1, min(days_n, 366))
    except Exception:
        days_n = 7

    floor = (datetime.now() - timedelta(days=days_n + 2)).strftime("%Y-%m-%d 00:00:00")

    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT date(REPLACE(created_at, 'T', ' ')) AS d, COUNT(*) AS c
            FROM messages
            WHERE network_id = ?
              AND is_valid = 1
              AND coalesce(content_type, 'intercept') = 'intercept'
              AND REPLACE(created_at, 'T', ' ') >= ?
            GROUP BY d
            ORDER BY d
            """,
            (nid, floor),
        ).fetchall()

    return {
        "ok": True,
        "days": days_n,
        "rows": [{"date": r["d"], "count": int(r["c"])} for r in rows if r["d"]],
    }


@router.get("/api/networks/{network_id}/peleng")
def api_network_peleng(network_id: int, days: int = 7):
    """Return peleng batches/points for a given network.

    Current DB schema stores `network_id` in `peleng_batches`, so we load
    batches directly by `network_id` and join frequency from `networks`.
    """

    try:
        nid = int(network_id)
    except Exception:
        return {"ok": False, "error": "invalid network_id"}

    try:
        days_n = int(days)
        if days_n < 1:
            days_n = 1
        if days_n > 365:
            days_n = 365
    except Exception:
        days_n = 7

    now = datetime.now().replace(microsecond=0)
    to_dt = now.isoformat(timespec="seconds").replace("T", " ")
    from_dt = (now - timedelta(days=days_n)).isoformat(timespec="seconds").replace("T", " ")

    with get_conn() as conn:
        net_row = conn.execute(
            "SELECT id, frequency, mask FROM networks WHERE id=? LIMIT 1",
            (nid,),
        ).fetchone()

        if not net_row:
            return {"ok": False, "error": "network not found"}

        batch_rows = conn.execute(
            """
            SELECT pb.id, pb.event_dt, n.frequency
            FROM peleng_batches pb
            JOIN networks n ON n.id = pb.network_id
            WHERE pb.network_id = ?
              AND pb.event_dt >= ?
              AND pb.event_dt <= ?
            ORDER BY pb.event_dt DESC, pb.id DESC
            LIMIT 80
            """,
            (nid, from_dt, to_dt),
        ).fetchall()

        batch_ids = [int(r["id"]) for r in batch_rows]
        points_rows = []
        if batch_ids:
            placeholders = ",".join(["?"] * len(batch_ids))
            points_rows = conn.execute(
                f"""
                SELECT batch_id, mgrs
                FROM peleng_points
                WHERE batch_id IN ({placeholders})
                ORDER BY batch_id ASC, id ASC
                """,
                tuple(batch_ids),
            ).fetchall()

    points_by_batch: dict[int, list[str]] = {}
    for r in points_rows:
        bid = int(r["batch_id"])
        mgrs = str(r["mgrs"] or "").strip()
        if not mgrs:
            continue
        points_by_batch.setdefault(bid, []).append(mgrs)

    batches_out = []
    points_total = 0
    for br in batch_rows:
        bid = int(br["id"])
        pts = points_by_batch.get(bid, [])
        points_total += len(pts)
        batches_out.append(
            {
                "id": bid,
                "event_dt": str(br["event_dt"] or ""),
                "frequency": str(br["frequency"] or ""),
                "points_count": len(pts),
                "points": pts[:200],  # cap to avoid huge payload
            }
        )

    # Keep payload size bounded: if many points exist in a batch, tell UI.
    for b in batches_out:
        full_cnt = len(points_by_batch.get(b["id"], []))
        b["points_truncated"] = full_cnt > b["points_count"] or b["points_count"] > len(b["points"])

    return {
        "ok": True,
        "batches": batches_out,
        "meta": {"days": days_n, "from_dt": from_dt, "to_dt": to_dt, "points_total": points_total},
    }


@router.get("/api/networks/lookup")
def api_networks_lookup(q: str):
    """Lookup networks matching a query for autocomplete/select widgets.

    Args:
        q: query string (frequency/mask/free text).

    Returns:
        dict: `{"ok": True, "rows": [...]}` where rows contain minimal
        network metadata.
    """

    query = (q or "").strip()
    if not query:
        return {"ok": True, "rows": []}

    with get_conn() as conn:
        rows = search_network_rows(conn, query, limit=60)

    out = []
    for r in rows:
        out.append({
            "id": int(r["id"]),
            "frequency": r["frequency"] or "",
            "mask": r["mask"] or "",
            "unit": r["unit"] or "",
        })

    return {"ok": True, "rows": out}


@router.get("/api/networks/by-id")
def api_network_by_id(id: int):
    """Fetch a single network by id for UI components.

    Args:
        id: network id.

    Returns:
        dict: `{"ok": True, "row": {...}}` or `{"ok": True, "row": None}`.
    """
    try:
        network_id = int(id)
    except Exception:
        return {"ok": False, "error": "invalid id"}

    with get_conn() as conn:
        row = conn.execute(
            "SELECT id, frequency, mask, unit FROM networks WHERE id=? LIMIT 1",
            (network_id,),
        ).fetchone()

        if not row:
            return {"ok": True, "row": None}

        return {
            "ok": True,
            "row": {
                "id": int(row["id"]),
                "frequency": row["frequency"] or "",
                "mask": row["mask"] or "",
                "unit": row["unit"] or "",
            },
        }


@router.get("/api/networks/{network_id}/cross-analysis")
def api_network_cross_analysis(
    network_id: int,
    active_only: int = 1,
    min_matches: int = 5,
):
    """Крос-аналіз спільних позивних у межах бригадно-полкової групи.

    Порівнюємо цільову р/м з усіма р/м тієї самої групи ("Хто (група)",
    networks.group_id). Якщо active_only=1 — лише з тими, що мають статус
    мережі «Спостерігається». Позивні зіставляються за нормалізованою назвою
    (callsigns.name; «НВ» виключаємо). До таблиці потрапляють лише ті р/м, що
    мають >= min_matches спільних позивних із цільовою.

    Відповідь:
        target  — {id, frequency, unit, group_name, status}
        peers   — усі р/м групи зі спільним лічильником (для контейнера-зведення)
        columns — peers, що пройшли min_matches (стовпці таблиці)
        rows    — позивні цільової р/м, наявні хоча б в одному column-peer:
                  {name, target:{...}, cells:{<net_id>:{callsign_id,status_id,status_label}}}
    """
    min_matches = max(1, int(min_matches or 1))
    with get_conn() as conn:
        tgt = conn.execute(
            "SELECT id, frequency, unit, group_id, status_id FROM networks WHERE id=?",
            (network_id,),
        ).fetchone()
        if not tgt:
            return JSONResponse({"ok": False, "error": "Радіомережу не знайдено"}, status_code=404)

        group_name = ""
        if tgt["group_id"]:
            g = conn.execute("SELECT name FROM groups WHERE id=?", (tgt["group_id"],)).fetchone()
            group_name = (g["name"] if g else "") or ""

        st = conn.execute("SELECT name FROM statuses WHERE id=?", (tgt["status_id"],)).fetchone()
        target_status = (st["name"] if st else "") or ""

        # Статус мережі «Спостерігається» (для фільтра «Активні р/м»).
        obs = conn.execute(
            "SELECT id FROM statuses WHERE name='Спостерігається' LIMIT 1"
        ).fetchone()
        obs_id = obs["id"] if obs else None

        # Назви статусів позивних (callsign_statuses) для підказок.
        cs_status_label = {
            int(r["id"]): (r["name"] or "")
            for r in conn.execute("SELECT id, name FROM callsign_statuses").fetchall()
        }

        # Позивні цільової р/м (без «НВ»).
        tgt_rows = conn.execute(
            "SELECT id, name, COALESCE(callsign_status_id, status_id) AS sid "
            "FROM callsigns WHERE network_id=? AND name <> 'НВ'",
            (network_id,),
        ).fetchall()
        target_cs = {
            r["name"]: {
                "callsign_id": int(r["id"]),
                "status_id": int(r["sid"]) if r["sid"] is not None else None,
            }
            for r in tgt_rows
        }
        target_names = list(target_cs.keys())

        # Р/м тієї самої групи (окрім цільової), опційно лише «Спостерігається».
        peer_wheres = ["group_id = ?", "id <> ?"]
        peer_params: list = [tgt["group_id"] or 0, network_id]
        if active_only and obs_id is not None:
            peer_wheres.append("status_id = ?")
            peer_params.append(obs_id)
        peer_rows = conn.execute(
            "SELECT id, frequency, unit, status_id FROM networks WHERE "
            + " AND ".join(peer_wheres),
            peer_params,
        ).fetchall()
        peers_meta = {
            int(r["id"]): {
                "network_id": int(r["id"]),
                "frequency": r["frequency"] or "",
                "unit": r["unit"] or "",
            }
            for r in peer_rows
        }

        # Спільні позивні по всіх peer-мережах одним запитом.
        # cells[name][net_id] = {callsign_id, status_id, status_label}
        cells: dict = {}
        common_by_net: dict = {pid: 0 for pid in peers_meta}
        if target_names and peers_meta:
            ph_nets = ",".join("?" * len(peers_meta))
            ph_names = ",".join("?" * len(target_names))
            rows = conn.execute(
                f"SELECT network_id, name, id, "
                f"       COALESCE(callsign_status_id, status_id) AS sid "
                f"FROM callsigns "
                f"WHERE network_id IN ({ph_nets}) AND name <> 'НВ' "
                f"  AND name IN ({ph_names})",
                [*peers_meta.keys(), *target_names],
            ).fetchall()
            for r in rows:
                nid = int(r["network_id"])
                nm = r["name"]
                sid = int(r["sid"]) if r["sid"] is not None else None
                cells.setdefault(nm, {})[nid] = {
                    "callsign_id": int(r["id"]),
                    "status_id": sid,
                    "status_label": cs_status_label.get(sid, "") if sid else "",
                }
                common_by_net[nid] = common_by_net.get(nid, 0) + 1

        # Зведення по всіх peer (для контейнера), відсортоване за спільними desc.
        peers_summary = sorted(
            (
                {**peers_meta[pid], "common": common_by_net.get(pid, 0)}
                for pid in peers_meta
            ),
            key=lambda p: (-p["common"], p["frequency"]),
        )
        # Стовпці таблиці — ті, що пройшли поріг.
        columns = [p for p in peers_summary if p["common"] >= min_matches]
        col_ids = [c["network_id"] for c in columns]

        # Рядки — позивні цільової, наявні хоча б в одному column-peer.
        out_rows = []
        for nm in target_names:
            row_cells = {
                nid: cells[nm][nid]
                for nid in col_ids
                if nm in cells and nid in cells[nm]
            }
            if not row_cells:
                continue
            tinfo = target_cs[nm]
            out_rows.append({
                "name": nm,
                "target": {
                    "callsign_id": tinfo["callsign_id"],
                    "status_id": tinfo["status_id"],
                    "status_label": cs_status_label.get(tinfo["status_id"], "") if tinfo["status_id"] else "",
                },
                "cells": row_cells,
                "_n": len(row_cells),
            })
        out_rows.sort(key=lambda r: (-r["_n"], r["name"]))
        for r in out_rows:
            r.pop("_n", None)

    return {
        "ok": True,
        "target": {
            "id": int(tgt["id"]),
            "frequency": tgt["frequency"] or "",
            "unit": tgt["unit"] or "",
            "group_name": group_name,
            "status": target_status,
            "total_callsigns": len(target_names),
        },
        "peers": peers_summary,
        "columns": columns,
        "rows": out_rows,
        "min_matches": min_matches,
        "active_only": bool(active_only),
    }


@router.get("/api/network-aliases")
def api_network_aliases_list(actor=Depends(get_actor)):
    """List active (non-archived) network aliases with linked network frequency."""
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT na.id, na.network_id, na.alias_text, n.frequency
            FROM network_aliases na
            JOIN networks n ON n.id = na.network_id
            WHERE COALESCE(na.is_archived, 0) = 0
            ORDER BY n.frequency COLLATE NOCASE, na.alias_text COLLATE NOCASE
            """
        ).fetchall()

    out = []
    for r in rows:
        out.append(
            {
                "id": int(r["id"]),
                "network_id": int(r["network_id"]),
                "alias_text": r["alias_text"] or "",
                "frequency": r["frequency"] or "",
            }
        )
    return {"ok": True, "rows": out}


@router.post("/api/network-aliases")
def api_network_aliases_create(body: NetworkAliasCreateIn, actor=Depends(get_actor)):
    """Create a new network alias row."""
    alias_text = _clean_alias_text(body.alias_text)
    if not alias_text.strip():
        return JSONResponse(
            {"ok": False, "error": "Аліас не може бути порожнім."},
            status_code=400,
        )

    try:
        nid = int(body.network_id)
    except Exception:
        return JSONResponse(
            {"ok": False, "error": "Некоректна радіомережа."},
            status_code=400,
        )

    new_id: Optional[int] = None
    with get_conn() as conn:
        row = conn.execute("SELECT id FROM networks WHERE id=? LIMIT 1", (nid,)).fetchone()
        if not row:
            return JSONResponse(
                {"ok": False, "error": "Радіомережу не знайдено."},
                status_code=404,
            )
        try:
            cur = conn.execute(
                "INSERT INTO network_aliases (network_id, alias_text, is_archived) VALUES (?,?,0)",
                (nid, alias_text),
            )
            new_id = int(cur.lastrowid)
        except sqlite3.IntegrityError:
            return JSONResponse(
                {
                    "ok": False,
                    "error": "Такий аліас уже існує або дублюється для цієї мережі.",
                },
                status_code=400,
            )

    return {"ok": True, "id": new_id}


@router.patch("/api/network-aliases/{alias_id}")
def api_network_aliases_update(
    alias_id: int,
    body: NetworkAliasUpdateIn,
    actor=Depends(get_actor),
):
    """Update alias text for an active alias row."""
    alias_text = _clean_alias_text(body.alias_text)
    if not alias_text.strip():
        return JSONResponse(
            {"ok": False, "error": "Аліас не може бути порожнім."},
            status_code=400,
        )

    try:
        aid = int(alias_id)
    except Exception:
        return JSONResponse(
            {"ok": False, "error": "Некоректний ідентифікатор."},
            status_code=400,
        )

    with get_conn() as conn:
        row = conn.execute(
            "SELECT id FROM network_aliases WHERE id=? AND COALESCE(is_archived,0)=0",
            (aid,),
        ).fetchone()
        if not row:
            return JSONResponse(
                {"ok": False, "error": "Запис не знайдено."},
                status_code=404,
            )
        target_network_id = body.network_id
        if target_network_id is not None:
            net_row = conn.execute(
                "SELECT id FROM networks WHERE id=? LIMIT 1",
                (int(target_network_id),),
            ).fetchone()
            if not net_row:
                return JSONResponse(
                    {"ok": False, "error": "Радіомережу не знайдено."},
                    status_code=404,
                )
        try:
            if target_network_id is None:
                conn.execute(
                    "UPDATE network_aliases SET alias_text=? WHERE id=?",
                    (alias_text, aid),
                )
            else:
                conn.execute(
                    "UPDATE network_aliases SET alias_text=?, network_id=? WHERE id=?",
                    (alias_text, int(target_network_id), aid),
                )
        except sqlite3.IntegrityError:
            return JSONResponse(
                {"ok": False, "error": "Такий аліас уже існує."},
                status_code=400,
            )

    return {"ok": True}


def _norm_mask_key(m: str) -> str:
    """Pad the decimal part of a mask to exactly 4 digits.

    Ensures that ``200.224`` and ``200.2240`` compare as equal — both
    represent the same radio-network mask prefix, just written differently.
    """
    if "." in m:
        left, right = m.split(".", 1)
        right = (right + "0000")[:4]
        return f"{left}.{right}"
    return m


def _normalize_verify_key(raw: str) -> str:
    """Return a canonical key for a user-entered frequency/mask string.

    Masks (100/200/300 prefix) are normalized and padded to 4 decimal places.
    Regular frequencies are normalized to DDD.DDDD form.
    Falls back to the trimmed raw value if normalization returns nothing.
    """
    freq, mask = normalize_freq_or_mask(raw.strip())
    if mask:
        return _norm_mask_key(mask.rstrip("%"))
    if freq:
        return freq
    return raw.strip()


@router.post("/api/networks/verify")
def api_networks_verify(body: NetworkVerifyIn):
    """Compare a user-supplied list of frequencies/masks against
    networks that have chat='Очерет' AND status='Спостерігається'.

    Returns two lists:
    - ``not_in_input``: DB networks not covered by the supplied list.
    - ``not_in_db``:    Input items that have no matching DB network.
    """
    raw_items = [l.strip() for l in (body.items or []) if l.strip()]

    with get_conn() as conn:
        db_rows = conn.execute(
            """
            SELECT n.id, n.frequency,
                   COALESCE(n.mask, '')  AS mask,
                   COALESCE(n.unit, '')  AS unit,
                   COALESCE(n.zone, '')  AS zone
            FROM networks n
            JOIN chats    c ON c.id = n.chat_id
            JOIN statuses s ON s.id = n.status_id
            WHERE c.name = 'Очерет' AND s.name = 'Спостерігається'
            ORDER BY n.frequency
            """
        ).fetchall()

    # Build lookup: canonical_key -> network dict index.
    # A network registers under its frequency AND (if set) its mask.
    db_networks: list[dict] = []
    db_key_to_idx: dict[str, int] = {}

    for row in db_rows:
        net = {
            "id":        int(row["id"]),
            "frequency": row["frequency"] or "",
            "mask":      row["mask"] or "",
            "unit":      row["unit"] or "",
            "zone":      row["zone"] or "",
        }
        idx = len(db_networks)
        db_networks.append(net)
        if net["frequency"] and net["frequency"] not in db_key_to_idx:
            db_key_to_idx[net["frequency"]] = idx
        if net["mask"]:
            # Normalize DB mask to 4-decimal key so it matches input like
            # "200.224" (→ "200.2240") against stored "200.2240" correctly.
            norm_mk = _norm_mask_key(net["mask"])
            if norm_mk not in db_key_to_idx:
                db_key_to_idx[norm_mk] = idx

    # Normalize each input line to a comparable key.
    input_pairs: list[tuple[str, str]] = [
        (raw, _normalize_verify_key(raw)) for raw in raw_items
    ]

    # DB networks matched by at least one input item.
    matched_db_ids: set[int] = set()
    for _, norm in input_pairs:
        if norm in db_key_to_idx:
            matched_db_ids.add(db_networks[db_key_to_idx[norm]]["id"])

    not_in_input = [n for n in db_networks if n["id"] not in matched_db_ids]
    not_in_db    = [raw for raw, norm in input_pairs if norm not in db_key_to_idx]

    return {
        "ok":           True,
        "db_count":     len(db_networks),
        "not_in_input": not_in_input,
        "not_in_db":    not_in_db,
    }


@router.get("/api/networks/ocheret-frequencies")
def api_networks_ocheret_frequencies(format: str = "json", with_mask: int = 0):
    """Return current active "Очерет"-watched network frequencies.

    Filter: `chats.name = 'Очерет'` AND `statuses.name = 'Спостерігається'`.
    Used by operator scripts on multiple PCs that need an always-up-to-date
    list without manually syncing a local text file.

    Args:
        format:   'json' (default) → {ok, count, frequencies:[{frequency,mask}]};
                  'text'           → plain text, one entry per line.
        with_mask: 1 → for text format, append mask after a tab (if present).

    Returns:
        JSONResponse or PlainTextResponse, depending on `format`.
    """
    from fastapi.responses import PlainTextResponse

    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT n.frequency, COALESCE(n.mask, '') AS mask
            FROM networks n
            JOIN chats    c ON c.id = n.chat_id
            JOIN statuses s ON s.id = n.status_id
            WHERE c.name = 'Очерет' AND s.name = 'Спостерігається'
            ORDER BY n.frequency
            """
        ).fetchall()

    items = [{"frequency": r["frequency"] or "", "mask": r["mask"] or ""} for r in rows]

    if (format or "").lower() == "text":
        lines: list[str] = []
        if int(with_mask or 0):
            # Одна колонка: частота і маска (якщо є) — кожна на окремому рядку.
            # Це збігається з виглядом модалки "Очерет" на /networks і з тим,
            # що очікують операторські скрипти (read line-by-line).
            for it in items:
                if it["frequency"]: lines.append(it["frequency"])
                if it["mask"]:      lines.append(it["mask"])
        else:
            lines = [it["frequency"] for it in items if it["frequency"]]
        return PlainTextResponse("\n".join(lines), media_type="text/plain; charset=utf-8")

    return JSONResponse({"ok": True, "count": len(items), "frequencies": items})


def _fetchall(conn, sql: str, params=()):
    """Fetch all rows for a query (small helper for this router)."""
    cur = conn.execute(sql, params)
    return cur.fetchall()


def _fetchone(conn, sql: str, params=()):
    """Fetch a single row for a query (small helper for this router)."""
    cur = conn.execute(sql, params)
    return cur.fetchone()


def _lookup(conn):
    """Load reference data used by the networks UI."""
    statuses = _fetchall(conn, "SELECT id, name FROM statuses ORDER BY name")
    chats = _fetchall(conn, "SELECT id, name FROM chats ORDER BY name")
    groups = _fetchall(conn, "SELECT id, name FROM groups ORDER BY name")
    tags = _fetchall(
        conn,
        "SELECT id, name, COALESCE(tag_group,'main') AS tag_group FROM network_tags ORDER BY name",
    )

    status_map = {r["id"]: r["name"] for r in statuses}
    chat_map = {r["id"]: r["name"] for r in chats}
    group_map = {r["id"]: r["name"] for r in groups}
    tag_map = {r["id"]: r["name"] for r in tags}

    return statuses, chats, groups, tags, status_map, chat_map, group_map, tag_map


def _get_tag_ids(conn, network_id: int) -> List[int]:
    """Return tag ids currently assigned to a network."""
    rows = _fetchall(conn, "SELECT tag_id FROM network_tag_links WHERE network_id=?", (network_id,))
    return [int(r["tag_id"]) for r in rows]


def _ensure_etalon(conn, network_id: int):
    """Ensure an `etalons` row exists for the given network."""
    row = _fetchone(conn, "SELECT id FROM etalons WHERE network_id=?", (network_id,))
    if row:
        return
    now = datetime.utcnow().isoformat(timespec="seconds")
    conn.execute(
        "INSERT INTO etalons(network_id, start_date, end_date, updated_at) VALUES (?,?,?,?)",
        (network_id, None, None, now),
    )
    

def _etalon_row_has_content(conn, network_id: int) -> bool:
    """True if the etalon row has any saved text or dates."""
    row = _fetchone(
        conn,
        """
        SELECT correspondents, start_date, end_date
        FROM etalons
        WHERE network_id = ?
        """,
        (int(network_id),),
    )
    if not row:
        return False

    def nz(col: str) -> bool:
        try:
            v = row[col]
        except Exception:
            return False
        return bool(v is not None and str(v).strip())

    if nz("correspondents"):
        return True
    try:
        if row["start_date"] or row["end_date"]:
            return True
    except Exception:
        pass
    return False


def _get_etalon_dates(conn, network_id: int) -> tuple[Optional[date], Optional[date]]:
    """Return (start_date, end_date) for a network, if set and parseable."""
    row = _fetchone(conn, "SELECT start_date, end_date FROM etalons WHERE network_id=?", (network_id,))
    if not row:
        return None, None

    def parse_d(v) -> Optional[date]:
        if not v:
            return None
        try:
            return date.fromisoformat(str(v))
        except Exception:
            return None

    return parse_d(row["start_date"]), parse_d(row["end_date"])
    

def _load_network_card(conn, network_id: int):
    """Load a network card: current network row + selected tags + etalon date."""
    current = _fetchone(conn, "SELECT * FROM networks WHERE id=?", (int(network_id),))
    if not current:
        return None, [], None

    selected_tags = _get_tag_ids(conn, int(current["id"]))
    _ensure_etalon(conn, int(current["id"]))
    start_date_val, end_date_val = _get_etalon_dates(conn, int(current["id"]))
    return current, selected_tags, (start_date_val, end_date_val)


def _build_networks_context(
    request: Request,
    actor,
    statuses,
    chats,
    groups,
    tags,
    status_map,
    chat_map,
    group_map,
    tag_map,
    *,
    q_query: str = "",
    current=None,
    matches=None,
    match_tags=None,
    selected_tags=None,
    start_date_val: Optional[date] = None,
    end_date_val: Optional[date] = None,
    message: str = "",
    all_rows=None,
    all_selected_statuses=None,
    all_selected_chats=None,
    all_selected_groups=None,
    active_tab: str = "card",
    etalon_exists: bool = False,
):
    """Build Jinja template context for the networks page."""
    main_tags = [t for t in (tags or []) if t["tag_group"] == "main"]
    comp_tags  = [t for t in (tags or []) if t["tag_group"] == "composition"]
    return dict(
        request=request,
        actor=actor,
        q_query=q_query,
        q_freq=current["frequency"] if current else "",
        q_mask=current["mask"] if current else "",
        matches=matches or [],
        match_tags=match_tags or {},
        current=current,
        statuses=statuses,
        chats=chats,
        groups=groups,
        tags=tags,
        main_tags=main_tags,
        comp_tags=comp_tags,
        status_map=status_map,
        chat_map=chat_map,
        group_map=group_map,
        tag_map=tag_map,
        selected_tags=selected_tags or [],
        start_date=start_date_val,
        end_date=end_date_val,
        message=message,
        all_rows=all_rows or [],
        all_selected_statuses=all_selected_statuses or [],
        all_selected_chats=all_selected_chats or [],
        all_selected_groups=all_selected_groups or [],
        active_tab=active_tab or "card",
        etalon_exists=bool(etalon_exists),
    )


def _get_tag_ids_for_networks(conn, network_ids: List[int]) -> dict[int, List[int]]:
    """Return mapping: network_id -> list[tag_id]."""
    net_ids = [int(x) for x in (network_ids or []) if x is not None]
    if not net_ids:
        return {}

    placeholders = ",".join(["?"] * len(net_ids))
    rows = conn.execute(
        f"""
        SELECT network_id, tag_id
        FROM network_tag_links
        WHERE network_id IN ({placeholders})
        """,
        tuple(net_ids),
    ).fetchall()

    out: dict[int, List[int]] = {}
    for r in rows:
        nid = int(r["network_id"])
        tid = int(r["tag_id"])
        out.setdefault(nid, []).append(tid)
    return out


@router.get("/networks", response_class=HTMLResponse)
def networks_page(
    request: Request,
    id: Optional[int] = None,
    pick: Optional[int] = None,
    actor=Depends(get_actor),
):
    """Render networks page (initial view or when selecting a network)."""
    network_id = pick or id

    with get_conn() as conn:
        statuses, chats, groups, tags, status_map, chat_map, group_map, tag_map = _lookup(conn)
        all_selected_statuses = _default_all_status_ids(conn)
        all_selected_chats = [c["id"] for c in chats]
        all_selected_groups = [g["id"] for g in groups]
        all_rows = _all_networks_list(conn, all_selected_statuses, all_selected_chats, all_selected_groups)

        current = None
        selected_tags: List[int] = []
        start_date_val: Optional[date] = None
        end_date_val: Optional[date] = None
        q_query = ""
        message = request.query_params.get("msg", "")
        draft = request.session.pop("network_save_draft", None)

        etalon_exists = False
        if network_id:
            current, selected_tags, dates = _load_network_card(conn, int(network_id))
            if dates:
                start_date_val, end_date_val = dates
            if current:
                q_query = current["frequency"] or current["mask"] or ""
                etalon_exists = _etalon_row_has_content(conn, int(current["id"]))
        elif draft:
            q_query = draft.get("frequency", "") or draft.get("mask", "")
            start_date_raw = draft.get("start_date_str", "")
            if start_date_raw:
                try:
                    start_date_val = date.fromisoformat(start_date_raw)
                except Exception:
                    start_date_val = None
            end_date_raw = draft.get("end_date_str", "")
            if end_date_raw:
                try:
                    end_date_val = date.fromisoformat(end_date_raw)
                except Exception:
                    end_date_val = None
            selected_tags = draft.get("tag_ids") or []

    tab_raw = (request.query_params.get("tab") or "card").lower()
    if tab_raw not in ("card", "all", "etalons", "akademik"):
        tab_raw = "card"

    context = _build_networks_context(
        request,
        actor,
        statuses,
        chats,
        groups,
        tags,
        status_map,
        chat_map,
        group_map,
        tag_map,
        q_query=q_query,
        current=current,
        matches=[],
        selected_tags=selected_tags,
        start_date_val=start_date_val,
        end_date_val=end_date_val,
        message=message,
        all_rows=all_rows,
        all_selected_statuses=all_selected_statuses,
        all_selected_chats=all_selected_chats,
        all_selected_groups=all_selected_groups,
        active_tab=tab_raw,
        etalon_exists=etalon_exists,
    )
    if draft and not current:
        context["draft"] = draft
    return request.app.state.templates.TemplateResponse("networks.html", context)


@router.post("/networks/search", response_class=HTMLResponse)
def networks_search(
    request: Request,
    query: str = Form(""),
    actor=Depends(get_actor),
):
    """Search networks and render results in the networks page template."""
    q_query = (query or "").strip()

    with get_conn() as conn:
        statuses, chats, groups, tags, status_map, chat_map, group_map, tag_map = _lookup(conn)
        all_selected_statuses = _default_all_status_ids(conn)
        all_selected_chats = [0]
        all_selected_groups = [0]
        all_rows = _all_networks_list(conn, all_selected_statuses, all_selected_chats, all_selected_groups)

        current = None
        matches = []
        selected_tags: List[int] = []
        start_date_val: Optional[date] = None
        end_date_val: Optional[date] = None
        message = ""
        etalon_exists = False

        if not q_query:
            message = "Введи маску або частоту."
        else:
            matches = search_network_rows(conn, q_query, limit=100)

            if len(matches) == 1:
                current, selected_tags, dates = _load_network_card(conn, int(matches[0]["id"]))
                if dates:
                    start_date_val, end_date_val = dates
                if current:
                    etalon_exists = _etalon_row_has_content(conn, int(current["id"]))
            elif len(matches) == 0:
                message = "Не знайдено. Заповни картку і тисни “Зберегти”."

        match_ids: List[int] = []
        for r in matches or []:
            try:
                if r and r["id"] is not None:
                    match_ids.append(int(r["id"]))
            except Exception:
                continue
        match_tags = _get_tag_ids_for_networks(conn, match_ids) if match_ids else {}

        context = _build_networks_context(
            request,
            actor,
            statuses,
            chats,
            groups,
            tags,
            status_map,
            chat_map,
            group_map,
            tag_map,
            q_query=q_query,
            current=current,
            matches=matches,
            match_tags=match_tags,
            selected_tags=selected_tags,
            start_date_val=start_date_val,
            end_date_val=end_date_val,
            message=message,
            all_rows=all_rows,
            all_selected_statuses=all_selected_statuses,
            all_selected_chats=all_selected_chats,
            all_selected_groups=all_selected_groups,
            active_tab="card",
            etalon_exists=etalon_exists,
        )
        return request.app.state.templates.TemplateResponse("networks.html", context)


@router.post("/networks/all", response_class=HTMLResponse)
def networks_all_tab(
    request: Request,
    status_ids: list[int] = Form(default=[]),
    chat_ids: list[int] = Form(default=[]),
    group_ids: list[int] = Form(default=[]),
    actor=Depends(get_actor),
):
    """Apply filters for the "Усі р/м" tab within `/networks` page."""
    with get_conn() as conn:
        statuses, chats, groups, tags, status_map, chat_map, group_map, tag_map = _lookup(conn)
        rows = _all_networks_list(conn, status_ids, chat_ids, group_ids)

    context = _build_networks_context(
        request,
        actor,
        statuses,
        chats,
        groups,
        tags,
        status_map,
        chat_map,
        group_map,
        tag_map,
        q_query="",
        current=None,
        matches=[],
        match_tags={},
        selected_tags=[],
        start_date_val=None,
        end_date_val=None,
        message="",
        all_rows=rows,
        all_selected_statuses=status_ids,
        all_selected_chats=chat_ids,
        all_selected_groups=group_ids,
        active_tab="all",
    )
    return request.app.state.templates.TemplateResponse("networks.html", context)


def _set_tags(conn, network_id: int, tag_ids: List[int]):
    """Replace network tags with the provided tag id list."""
    conn.execute("DELETE FROM network_tag_links WHERE network_id=?", (network_id,))
    for tid in sorted(set(int(x) for x in tag_ids or [])):
        conn.execute(
            "INSERT OR IGNORE INTO network_tag_links(network_id, tag_id) VALUES (?,?)",
            (network_id, tid),
        )
        
        
def _set_etalon_dates(conn, network_id: int, start_date_str: str, end_date_str: str):
    """Update etalon start/end date fields for a network."""
    now = datetime.utcnow().isoformat(timespec="seconds")
    conn.execute(
        "UPDATE etalons SET start_date=?, end_date=?, updated_at=? WHERE network_id=?",
        (start_date_str or None, end_date_str or None, now, network_id),
    )


def _store_network_draft(request: Request, **fields):
    """Store last network form values in session for restore after validation errors."""
    request.session["network_save_draft"] = fields


def _draft_payload(
    frequency: str,
    mask: str,
    unit: str,
    zone: str,
    chat_id: int,
    group_id: int,
    status_id: int,
    comment: str,
    tag_ids: List[int],
    start_date_str: str,
    end_date_str: str,
):
    return {
        "frequency": frequency,
        "mask": mask,
        "unit": unit,
        "zone": zone,
        "chat_id": str(chat_id),
        "group_id": str(group_id),
        "status_id": str(status_id),
        "comment": comment,
        "tag_ids": [int(x) for x in tag_ids or []],
        "start_date_str": start_date_str,
        "end_date_str": end_date_str,
    }


def _missing_fields_message(
    *,
    frequency_ok: bool,
    unit: str,
    zone: str,
    chat_id: int,
    group_id: int,
    status_id: int,
) -> tuple[str, list[str]]:
    missing_labels: list[str] = []
    missing_keys: list[str] = []
    if not frequency_ok:
        missing_labels.append("частоту")
        missing_keys.append("frequency")
    if not (unit or "").strip():
        missing_labels.append("підрозділ")
        missing_keys.append("unit")
    if not (zone or "").strip():
        missing_labels.append("зону функціонування")
        missing_keys.append("zone")
    if not chat_id:
        missing_labels.append("чат (джерело)")
        missing_keys.append("chat_id")
    if not group_id:
        missing_labels.append("групу")
        missing_keys.append("group_id")
    if not status_id:
        missing_labels.append("статус")
        missing_keys.append("status_id")

    if not missing_labels:
        return "", []
    if len(missing_labels) == 1:
        return f"Потрібно вказати {missing_labels[0]}.", missing_keys
    if len(missing_labels) == 2:
        return f"Потрібно вказати {missing_labels[0]} та {missing_labels[1]}.", missing_keys
    return "Потрібно вказати: " + ", ".join(missing_labels[:-1]) + f" та {missing_labels[-1]}.", missing_keys
    
    
# ---------------------------------------------------------------------------
# Google Sheets sync (via Apps Script web-app — no credentials needed)
# ---------------------------------------------------------------------------

# URL is stored in config.env as SHEETS_SCRIPT_URL (not committed to git).


def _build_sheets_data(conn) -> list[list[str]]:
    """Build a 2-D list (rows × columns) to write to Google Sheets.

    Layout:
      Row 0  : headers  – «Хор», then one column per tag (alphabetical)
      Row 1+ : frequencies/masks, one entry per cell
    Column «Хор»  : all Спостерігається networks (freq + mask if both present)
    Tag columns   : Спостерігається networks that carry the tag
    """
    # All Спостерігається networks with their tags
    rows = conn.execute(
        """
        SELECT n.id, n.frequency, n.mask, nt.name AS tag_name
        FROM   networks n
        LEFT JOIN network_tag_links ntl ON ntl.network_id = n.id
        LEFT JOIN network_tags      nt  ON nt.id = ntl.tag_id
        WHERE  n.status_id = (SELECT id FROM statuses WHERE name = 'Спостерігається' LIMIT 1)
        ORDER  BY n.id
        """
    ).fetchall()

    # Gather all tag names in sorted order
    all_tag_names: list[str] = sorted({
        r["tag_name"] for r in rows if r["tag_name"]
    })

    # Build set of freq/mask entries per bucket (set to avoid duplicates)
    def _entries(freq, mask) -> list[str]:
        out = []
        if freq:
            out.append(freq)
        if mask and mask != freq:
            out.append(mask)
        return out

    # Хор bucket: unique entries across all Спостерігається networks
    seen_choir: set[str] = set()
    choir: list[str] = []
    # tag buckets
    tag_buckets: dict[str, list[str]] = {t: [] for t in all_tag_names}
    seen_tag: dict[str, set[str]] = {t: set() for t in all_tag_names}

    # Process rows (each row may repeat network_id when a network has multiple tags)
    processed_networks_choir: set[int] = set()
    processed_per_tag: dict[str, set[int]] = {t: set() for t in all_tag_names}

    for r in rows:
        nid = int(r["id"])
        freq = r["frequency"] or ""
        mask = r["mask"] or ""
        tag  = r["tag_name"]

        # Add to choir once per network
        if nid not in processed_networks_choir:
            processed_networks_choir.add(nid)
            for e in _entries(freq, mask):
                if e not in seen_choir:
                    seen_choir.add(e)
                    choir.append(e)

        # Add to tag bucket once per (network, tag) pair
        if tag and tag in tag_buckets:
            if nid not in processed_per_tag[tag]:
                processed_per_tag[tag].add(nid)
                for e in _entries(freq, mask):
                    if e not in seen_tag[tag]:
                        seen_tag[tag].add(e)
                        tag_buckets[tag].append(e)

    # Build 2-D grid
    headers = ["Хор"] + all_tag_names
    columns = [choir] + [tag_buckets[t] for t in all_tag_names]
    max_rows = max((len(c) for c in columns), default=0)

    grid: list[list[str]] = [headers]
    for i in range(max_rows):
        grid.append([col[i] if i < len(col) else "" for col in columns])

    return grid


def _sync_gsheets_bg(conn_factory):
    """Fire-and-forget: POST sheet data to the Apps Script web-app URL.

    Configure by placing the deployment URL (one line) in:
        secrets/sheets_script_url.txt
    No credentials or Google Cloud setup required.
    """
    def _run():
        script_url = settings.sheets_script_url.strip()
        if not script_url:
            log.debug("Google Sheets sync skipped: SHEETS_SCRIPT_URL not configured")
            return
        try:
            import json as _json
            import urllib.request as _req

            with conn_factory() as conn:
                data = _build_sheets_data(conn)

            payload = _json.dumps(data, ensure_ascii=False).encode()
            request = _req.Request(
                script_url,
                data=payload,
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            with _req.urlopen(request, timeout=20) as resp:
                body = resp.read().decode()
            log.info("Google Sheets sync OK: %s", body[:120])
        except Exception as exc:           # noqa: BLE001
            log.error("Google Sheets sync failed: %s", exc)

    threading.Thread(target=_run, daemon=True).start()


@router.post("/networks/save")
def networks_save(
    request: Request,
    frequency: str = Form(...),
    mask: str = Form(""),
    unit: str = Form(""),
    zone: str = Form(""),
    chat_id: int = Form(...),
    group_id: int = Form(...),
    status_id: int = Form(...),
    comment: str = Form(""),
    tag_ids: List[int] = Form(default=[]),
    start_date_str: str = Form(""),
    end_date_str: str = Form(""),
    actor=Depends(get_actor),
):
    """Insert or update a network record from the networks page form."""
    freq_norm = normalize_freq(frequency)

    # Normalize mask using the same rules as ingest/network_service so that
    # values entered on the networks page match tokens parsed from intercepts.
    raw_mask = (mask or "").strip()
    if raw_mask:
        mask_freq, mask_mask = normalize_freq_or_mask(raw_mask)
        # Prefer "mask-like" normalization; fall back to normalized frequency.
        if mask_mask:
            # Store mask in DB without trailing '%' so that users
            # don't see SQL wildcard characters in the UI. The '%'
            # is only needed in query patterns, not in persisted values.
            mask_val = mask_mask[:-1] if mask_mask.endswith("%") else mask_mask
        else:
            mask_val = mask_freq
    else:
        mask_val = None
    unit_val = (unit or "").strip() or None
    zone_val = (zone or "").strip() or None
    comment_val = (comment or "").strip() or None

    message, missing_keys = _missing_fields_message(
        frequency_ok=bool(freq_norm),
        unit=unit_val or "",
        zone=zone_val or "",
        chat_id=chat_id,
        group_id=group_id,
        status_id=status_id,
    )
    if message:
        _store_network_draft(
            request,
            **{
                **_draft_payload(
                frequency=frequency,
                mask=mask,
                unit=unit,
                zone=zone,
                chat_id=chat_id,
                group_id=group_id,
                status_id=status_id,
                comment=comment,
                tag_ids=tag_ids,
                start_date_str=start_date_str,
                end_date_str=end_date_str,
                ),
                "missing_fields": missing_keys,
            },
        )
        return RedirectResponse(url=f"/networks?msg={message}", status_code=303)

    now = datetime.utcnow().isoformat(timespec="seconds")

    with get_conn() as conn:
        existing = _fetchone(conn, "SELECT id FROM networks WHERE frequency=?", (freq_norm,))

        if not existing:
            cur = conn.execute(
                """
                INSERT INTO networks (frequency, mask, unit, zone, chat_id, group_id, status_id, comment, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?)
                """,
                (freq_norm, mask_val, unit_val, zone_val, int(chat_id), int(group_id), int(status_id), comment_val, now),
            )
            network_id = int(cur.lastrowid)
        else:
            network_id = int(existing["id"])
            conn.execute(
                """
                UPDATE networks
                SET mask=?, unit=?, zone=?, chat_id=?, group_id=?, status_id=?, comment=?, updated_at=?
                WHERE id=?
                """,
                (mask_val, unit_val, zone_val, int(chat_id), int(group_id), int(status_id), comment_val, now, network_id),
            )

        _set_tags(conn, network_id, tag_ids)

        _ensure_etalon(conn, network_id)
        if start_date_str or end_date_str:
            _set_etalon_dates(conn, network_id, start_date_str, end_date_str)

        # When a network is marked "Мертва" it will no longer receive new
        # intercepts, so the callsign correction map becomes irrelevant.
        # Clean it up to avoid stale data.
        dead_status = conn.execute(
            "SELECT id FROM statuses WHERE name = 'Мертва' LIMIT 1"
        ).fetchone()
        if dead_status and int(status_id) == int(dead_status["id"]):
            conn.execute(
                "DELETE FROM callsign_corrections WHERE network_id = ?",
                (network_id,),
            )

        conn.commit()

    # Sync to Google Sheets in background (non-blocking)
    _sync_gsheets_bg(get_conn)

    request.session.pop("network_save_draft", None)
    return RedirectResponse(url=f"/networks?pick={network_id}", status_code=303)