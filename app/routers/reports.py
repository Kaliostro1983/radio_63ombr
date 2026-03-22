from __future__ import annotations

from datetime import datetime
from io import BytesIO
import csv
import urllib.request
import logging
from urllib.parse import quote

from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import RedirectResponse, StreamingResponse

from app.core.db import get_conn
from app.core.normalize import normalize_freq_or_mask
from app.reports.enemy_moves_config import load_reports_config
from app.reports.enemy_moves_report import EnemyMoveItem, build_enemy_moves_docx_bytes

from openpyxl import load_workbook


router = APIRouter(tags=["reports"])

def _ascii_filename_fallback(filename: str) -> str:
    """Return a latin-1 safe fallback filename (best-effort)."""
    safe = (filename or "").strip()
    if not safe:
        return "report.docx"
    out: list[str] = []
    for ch in safe:
        o = ord(ch)
        if 32 <= o < 127 and ch not in {'"', "\\"}:
            out.append(ch)
        elif ch in {" ", "-", "_", ".", "(", ")", "[", "]"}:
            out.append(ch)
        else:
            out.append("_")
    fallback = "".join(out).strip() or "report.docx"
    if not fallback.lower().endswith(".docx"):
        fallback += ".docx"
    return fallback


def _content_disposition_attachment(filename: str) -> str:
    """RFC 6266 / RFC 5987 compatible Content-Disposition for UTF-8 filenames."""
    fallback = _ascii_filename_fallback(filename)
    utf8 = quote(filename, safe="")
    return f"attachment; filename=\"{fallback}\"; filename*=UTF-8''{utf8}"


@router.get("/reports")
def reports_page_redirect():
    """Звіти перенесено на Головну → вкладка «Звіти»."""
    return RedirectResponse(url="/home?tab=reports", status_code=302)

def _load_moves_from_xlsx_bytes(content: bytes) -> list[tuple[str, str]]:
    """Return list of (freq_or_mask, move_text) from the first worksheet."""
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    # Find headers in first row
    headers: dict[str, int] = {}
    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1):
        if not cell:
            continue
        name = str(cell).strip()
        headers[name] = idx

    col_freq = headers.get("Частота/маска")
    col_move = headers.get("Переміщення")
    if not col_freq or not col_move:
        raise ValueError("У XLSX мають бути колонки 'Частота/маска' та 'Переміщення'.")

    result: list[tuple[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_freq = row[col_freq - 1] if len(row) >= col_freq else None
        raw_move = row[col_move - 1] if len(row) >= col_move else None
        freq = str(raw_freq).strip() if raw_freq is not None else ""
        move = str(raw_move).strip() if raw_move is not None else ""
        if not freq or not move:
            continue
        result.append((freq, move))
    return result


def _load_moves_from_gsheet_csv(spreadsheet_id: str, gid: str | None) -> list[tuple[str, str]]:
    """Load moves from Google Sheets CSV export.

    Requires the sheet to be accessible without auth in the current environment.
    """
    sid = (spreadsheet_id or "").strip()
    if not sid:
        raise ValueError("Не задано spreadsheet_id для gsheet_csv.")

    url = f"https://docs.google.com/spreadsheets/d/{sid}/export?format=csv"
    gid_val = (gid or "").strip()
    if gid_val:
        url += f"&gid={gid_val}"

    req = urllib.request.Request(url, headers={"User-Agent": "radio_63ombr/1.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        raw = resp.read()

    # Google may include UTF-8 BOM
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(text.splitlines())

    # Be tolerant to minor header whitespace
    fieldnames = [fn.strip() for fn in (reader.fieldnames or [])]
    # Map normalized header -> original
    header_map = {fn.strip(): fn for fn in (reader.fieldnames or [])}
    col_freq = header_map.get("Частота/маска")
    col_move = header_map.get("Переміщення")
    if not col_freq or not col_move:
        raise ValueError("У Google Sheet мають бути колонки 'Частота/маска' та 'Переміщення'.")

    result: list[tuple[str, str]] = []
    for row in reader:
        freq = str(row.get(col_freq) or "").strip()
        move = str(row.get(col_move) or "").strip()
        if not freq or not move:
            continue
        result.append((freq, move))
    return result


@router.post("/reports/enemy-moves")
def generate_enemy_moves_report(moves_xlsx: UploadFile | None = File(default=None)):
    """Generate enemies moves report from uploaded XLSX or configured Google Sheets CSV."""
    try:
        moves: list[tuple[str, str]]
        if moves_xlsx is not None and getattr(moves_xlsx, "filename", ""):
            content = moves_xlsx.file.read()
            moves = _load_moves_from_xlsx_bytes(content)
        else:
            cfg = load_reports_config()
            moves_cfg = cfg.get("moves") or {}
            if not isinstance(moves_cfg, dict):
                moves_cfg = {}
            source = str(moves_cfg.get("source") or "gsheet_csv").strip()
            if source == "xlsx":
                raise ValueError("Джерело moves=xlsx у веб-версії підтримується через завантаження файлу.")
            spreadsheet_id = str(moves_cfg.get("spreadsheet_id") or "").strip()
            gid = str(moves_cfg.get("gid") or "").strip()
            moves = _load_moves_from_gsheet_csv(spreadsheet_id, gid or None)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    items: list[EnemyMoveItem] = []
    missing_network: set[str] = set()
    missing_group: set[str] = set()

    with get_conn() as conn:
        for raw_freq_or_mask, move_text in moves:
            exact, mask = normalize_freq_or_mask(raw_freq_or_mask)

            row = None
            if mask:
                # Use LIKE against frequency and mask fields.
                row = conn.execute(
                    """
                    SELECT n.unit AS unit, g.name AS group_name, n.frequency AS frequency
                    FROM networks n
                    LEFT JOIN groups g ON g.id = n.group_id
                    WHERE (n.frequency LIKE ? OR n.mask LIKE ?)
                    ORDER BY n.frequency ASC
                    LIMIT 1
                    """,
                    (mask, mask),
                ).fetchone()
            elif exact:
                row = conn.execute(
                    """
                    SELECT n.unit AS unit, g.name AS group_name, n.frequency AS frequency
                    FROM networks n
                    LEFT JOIN groups g ON g.id = n.group_id
                    WHERE n.frequency = ? OR n.mask = ?
                    LIMIT 1
                    """,
                    (exact, exact),
                ).fetchone()

            if not row:
                missing_network.add(str(raw_freq_or_mask))
                continue

            group = (row["group_name"] or "").strip()
            if not group:
                missing_group.add(str(raw_freq_or_mask))
                continue

            unit = (row["unit"] or "").strip()
            freq_display = (exact or row["frequency"] or raw_freq_or_mask or "").strip()

            items.append(
                EnemyMoveItem(
                    freq=freq_display,
                    move=move_text,
                    unit=unit,
                    group=group,
                )
            )

    if missing_network:
        sample = sorted(missing_network)[:50]
        logging.warning(
            "Enemy moves report: не знайдено радіомережу в БД для %d рядків. Частоти/маски (перші %d): %s",
            len(missing_network),
            len(sample),
            ", ".join(sample),
        )

    if missing_group:
        sample = sorted(missing_group)[:50]
        logging.warning(
            "Enemy moves report: знайдено мережу, але відсутня group_id/назва групи для %d рядків. Частоти/маски (перші %d): %s",
            len(missing_group),
            len(sample),
            ", ".join(sample),
        )

    content, filename = build_enemy_moves_docx_bytes(items, report_date=datetime.now().date())
    bio = BytesIO(content)
    headers = {"Content-Disposition": _content_disposition_attachment(filename)}
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers=headers,
    )

