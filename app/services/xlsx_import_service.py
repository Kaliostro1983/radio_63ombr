"""XLSX import service for intercepts.

This module implements importing intercept messages from an Excel file.

Usage in the system:

- The web UI allows uploading an `.xlsx` file.
- The router forwards the file path (or uploaded content) to this service.
- The service reads the workbook, extracts message text from a target
  column, and feeds each row into the same ingest pipeline
  (`process_whatsapp_payload`) as real-time sources.

The import result is a summary dict containing counts for processed rows,
inserted messages, duplicates, failures, and skips.
"""

from __future__ import annotations

from typing import Dict, Any
from openpyxl import load_workbook
from pathlib import Path
from uuid import uuid4

from app.services.ingest_service import process_whatsapp_payload
from app.core.logging import get_logger


log = get_logger("xlsx_import_service")


TARGET_COLUMN = "р\\обмін"


def _normalize_header(value) -> str:
    """Normalize an XLSX header cell into a comparable string."""
    if value is None:
        return ""
    return str(value).strip().lower()


def _normalize_cell_text(value: Any) -> str:
    """Normalize an XLSX cell value into clean text for parsing.

    The function:
    - normalizes Windows-style line endings;
    - removes common invisible Unicode characters that break parsers.

    Args:
        value: raw cell value.

    Returns:
        str: cleaned text (may be empty).
    """
    if value is None:
        return ""

    text = str(value)

    # Excel/Windows line endings -> unix
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    # прибираємо сміттєві unicode-символи, які часто ламають парсер
    text = text.replace("\u00a0", " ")   # NBSP
    text = text.replace("\u200b", "")    # zero-width space
    text = text.replace("\ufeff", "")    # BOM

    return text.strip()


def _fmt_date(value: Any) -> str:
    """\u0414\u0430\u0442\u0430 \u043a\u043e\u043c\u0456\u0440\u043a\u0438 \u2192 'DD.MM.YYYY' (\u043f\u0440\u0438\u0439\u043c\u0430\u0454 datetime/date/\u0440\u044f\u0434\u043e\u043a)."""
    import datetime as _dt
    if isinstance(value, _dt.datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, _dt.date):
        return value.strftime("%d.%m.%Y")
    s = str(value or "").strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S"):
        try:
            return _dt.datetime.strptime(s[:19], fmt).strftime("%d.%m.%Y")
        except Exception:
            pass
    return s


def _assemble_intercept_text(row, ci, body: str) -> str:
    """\u0421\u043a\u043b\u0430\u0441\u0442\u0438 \u0442\u0435\u043a\u0441\u0442 \u043f\u0435\u0440\u0435\u0445\u043e\u043f\u043b\u0435\u043d\u043d\u044f \u0437\u0430 \u0448\u0430\u0431\u043b\u043e\u043d\u043e\u043c \u0456\u0437 \u043e\u043a\u0440\u0435\u043c\u0438\u0445 \u043a\u043e\u043b\u043e\u043d\u043e\u043a (\u043d\u043e\u0432\u0438\u0439 \u0444\u043e\u0440\u043c\u0430\u0442).

    \u0428\u0430\u0431\u043b\u043e\u043d:
        DD.MM.YYYY HH:MM:SS
        \u0447\u0430\u0441\u0442\u043e\u0442\u0430
        \u043d\u0430\u0437\u0432\u0430 \u0440/\u043c
        \u0445\u0442\u043e
        \u043a\u043e\u043c\u0443
        <\u043f\u043e\u0440\u043e\u0436\u043d\u0456\u0439 \u0440\u044f\u0434\u043e\u043a>
        \u0442\u0456\u043b\u043e (\u0440\\\u043e\u0431\u043c\u0456\u043d)
    """
    def g(name: str) -> str:
        i = ci.get(name)
        return _normalize_cell_text(row[i]) if (i is not None and i < len(row)) else ""

    di = ci.get("\u0434\u0430\u0442\u0430")
    date = _fmt_date(row[di]) if (di is not None and di < len(row)) else ""
    dt = (date + " " + g("\u0447\u0430\u0441")).strip()
    who = g("\u0445\u0442\u043e") or "\u041d\u0412"
    whom = g("\u043a\u043e\u043c\u0443") or "\u041d\u0412"
    return "\n".join([dt, g("\u0447\u0430\u0441\u0442\u043e\u0442\u0430"), g("\u043d\u0430\u0437\u0432\u0430 \u0440\\\u043c"), who, whom, "", body])


def import_xlsx(file_path: str, progress_cb=None) -> Dict[str, Any]:
    """Import intercept messages from an XLSX file.

    Args:
        file_path: path to `.xlsx` file.
        progress_cb: optional callable(stage: str, current: int, total: int)
            called periodically for UI progress.

    Returns:
        Dict[str, Any]: summary metrics including:
        `total_rows`, `processed`, `inserted`, `duplicates`, `failed`, `skipped`.

    Raises:
        ValueError: if the file is empty or target column is missing.
    """
    import_session_id = uuid4().hex

    def _report(stage: str, current: int, total: int) -> None:
        if progress_cb:
            try:
                progress_cb(stage, current, total)
            except Exception:
                pass

    _report("Читання файлу", 0, 0)

    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    # Загальна к-сть рядків (мінус заголовок) — для прогрес-лінії. У read_only
    # max_row може бути None → тоді прогрес без відсотка (лише лічильник).
    try:
        total_rows_est = max(0, int(sheet.max_row or 0) - 1)
    except Exception:
        total_rows_est = 0
    rows = sheet.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        raise ValueError("XLSX file is empty")

    headers = [_normalize_header(x) for x in header]

    # Індекси колонок за нормалізованими заголовками.
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h and h not in col_map:
            col_map[h] = i

    col_body = col_map.get(TARGET_COLUMN)  # 'р\обмін' — тіло діалогу
    if col_body is None:
        raise ValueError("Column 'р\\обмін' not found")

    # Новий (багатоколонковий) формат: шапку перехоплення складаємо з колонок
    # (Дата/Час/Частота/Назва р\\м/хто/кому), тіло — з 'р\\обмін'. Старий формат:
    # 'р\\обмін' уже містить увесь готовий текст перехоплення.
    multi_col = ("дата" in col_map) and ("частота" in col_map)

    col_index = col_body
    _report("Обробка рядків", 0, total_rows_est)

    summary = {
        "total_rows": 0,
        "processed": 0,
        "inserted": 0,
        "duplicates": 0,
        "failed": 0,
        "skipped": 0,
        "reasons": {},
        "reason_samples": {},
    }

    skip_reasons: dict[str, int] = {}
    reason_samples: dict[str, set[str]] = {}
    missing_network_samples: set[str] = set()
    net_line_invalid_count: int = 0

    filename = Path(file_path).name

    def _sample_text(row_no: int, text: str) -> str:
        # Keep row number so user can locate it in XLSX.
        raw = str(text or "")
        # Keep newlines (do not glue), but cap size.
        lines = raw.splitlines()
        head = lines[:8]
        clipped = "\n".join(head)
        if len(lines) > 8:
            clipped += "\n…"
        if len(clipped) > 500:
            clipped = clipped[:500].rstrip() + "…"
        return f"рядок {row_no}:\n{clipped}"

    for row_number, row in enumerate(rows, start=2):
        summary["total_rows"] += 1

        # Прогрес — не частіше ніж кожні 20 рядків (щоб не бити реєстр щоразу).
        if summary["total_rows"] % 20 == 0:
            _report("Обробка рядків", summary["total_rows"], total_rows_est)

        if col_index >= len(row):
            summary["skipped"] += 1
            reason = "xlsx_row_missing_target_column"
            skip_reasons[reason] = skip_reasons.get(reason, 0) + 1
            continue

        body = _normalize_cell_text(row[col_body])
        if not body:
            summary["skipped"] += 1
            reason = "xlsx_empty_cell"
            skip_reasons[reason] = skip_reasons.get(reason, 0) + 1
            continue

        # Новий формат — складаємо текст із колонок; старий — беремо як є.
        text = _assemble_intercept_text(row, col_map, body) if multi_col else body

        payload = {
            "platform": "xlsx",
            "chat_id": "xlsx_import",
            "chat_name": filename,
            "message_id": f"{import_session_id}:{filename}:{row_number}",
            "text": text,
        }

        try:
            result = process_whatsapp_payload(payload)
            summary["processed"] += 1

            if not result.get("ok", False):
                summary["failed"] += 1
                continue

            if result.get("duplicate"):
                summary["duplicates"] += 1
                continue

            if result.get("message_row_id"):
                summary["inserted"] += 1
                continue

            if result.get("peleng_batch_id"):
                summary["inserted"] += 1
                continue

            summary["skipped"] += 1
            reason = str(result.get("reason") or "skipped")
            skip_reasons[reason] = skip_reasons.get(reason, 0) + 1
            if reason == "net_line_invalid":
                net_line_invalid_count += 1
            elif len(reason_samples.get(reason, set())) < 5:
                s = reason_samples.setdefault(reason, set())
                s.add(_sample_text(row_number, text))
            if reason == "network_not_found":
                details = result.get("details") or {}
                freq = str(details.get("frequency") or "").strip()
                mask = str(details.get("mask") or "").strip()
                token = freq or mask or text[:80]
                if token:
                    missing_network_samples.add(token)

        except Exception:
            summary["failed"] += 1

    summary["reasons"] = skip_reasons
    summary["reason_samples"] = {
        k: sorted(v)[:5]
        for k, v in reason_samples.items()
        if k not in ("duplicate_message", "duplicate", "duplicates")
    }

    if missing_network_samples:
        sample = sorted(missing_network_samples)[:50]
        log.warning(
            "XLSX import: network_not_found for %d unique tokens. Samples: %s",
            len(missing_network_samples),
            ", ".join(sample),
        )

    # net_line_invalid is a parsing-quality signal; don't spam sample lines.
    # Print exactly one INFO message at the end (if present).
    if net_line_invalid_count > 0:
        log.info(
            "XLSX import: net_line_invalid occurred %d times (parsing degraded, frequency matching was still attempted).",
            net_line_invalid_count,
        )

    if skip_reasons:
        # Use NOTICE so it shows up even when INFO is filtered out.
        log.notice(
            "XLSX import skip reasons breakdown (%d kinds):",
            len(skip_reasons),
        )
        for k, v in sorted(skip_reasons.items(), key=lambda kv: (-kv[1], kv[0])):
            if k == "net_line_invalid":
                # Already reported once as INFO above.
                log.notice("  • %s: count=%d", k, v)
                continue

            samples = sorted(reason_samples.get(k, set()))
            log.notice("  • %s: count=%d", k, v)
            for i, s in enumerate(samples[:5], start=1):
                log.notice("      sample %d: %s", i, s)

    log.notice(
        "XLSX import done: total=%d processed=%d inserted=%d duplicates=%d skipped=%d failed=%d",
        summary["total_rows"],
        summary["processed"],
        summary["inserted"],
        summary["duplicates"],
        summary["skipped"],
        summary["failed"],
    )
    _report("Готово", summary["total_rows"], summary["total_rows"])
    return summary