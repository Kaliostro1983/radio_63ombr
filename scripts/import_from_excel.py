from __future__ import annotations
import os
import re
from datetime import datetime, date
from typing import Optional, Dict, Any, List

from openpyxl import load_workbook

from app.core.config import settings
from app.core.normalize import normalize_freq
from app.core.db import get_conn, init_db

CHAT_VALUES = {"Очерет", "Галявина", "ЦВО", "ОРК-ФМ", "Каменярі"}
SKIP_SHEETS = {"freq", "masks", "settings"}

def _parse_date_any(s: str) -> Optional[date]:
    s = s.strip()
    if not s:
        return None
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            return None
    m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except Exception:
            return None
    return None

def _sheet_name_is_freq(name: str) -> Optional[str]:
    name = str(name).strip()
    if re.fullmatch(r"\d{1,3}\.\d{1,4}", name):
        return normalize_freq(name)
    return None

def _parse_etalon_sheet(wb, sheet_name: str) -> Dict[str, Any]:
    ws = wb[sheet_name]
    lines: List[str] = []
    for row in ws.iter_rows(values_only=True):
        parts = []
        for v in row:
            if v is None:
                continue
            sv = str(v).strip()
            if sv:
                parts.append(sv)
        if parts:
            lines.append(" | ".join(parts))
    raw = "\n".join(lines)
    extracted: Dict[str, Any] = {"raw_import_text": raw}

    def find_item(n: int) -> Optional[str]:
        pat = re.compile(rf"^(?:{n}[\.|\)|:])\s*(.*)$")
        for ln in lines:
            m = pat.match(ln)
            if m:
                val = m.group(1).strip()
                if val:
                    return val
        return None

    def find_item_table(n: int) -> Optional[str]:
        for ln in lines:
            parts = [p.strip() for p in ln.split("|")]
            if not parts:
                continue
            first = parts[0]
            if re.fullmatch(rf"{n}[\.|\)]?", first):
                for cand in reversed(parts[1:]):
                    if cand:
                        return cand
        return None

    def get_item(n: int) -> Optional[str]:
        return find_item(n) or find_item_table(n)

    for key, n in [("purpose",3),("correspondents",4),("callsigns",5),("operation_mode",8),("traffic_type",9)]:
        val = get_item(n)
        if val:
            extracted[key] = val

    period = get_item(10)
    if period:
        d = _parse_date_any(period)
        if d:
            extracted["start_date"] = d
    if "start_date" not in extracted:
        d = _parse_date_any(raw)
        if d:
            extracted["start_date"] = d
    return extracted

def _get_id_map(conn, table: str) -> Dict[str, int]:
    cur = conn.execute(f"SELECT id, name FROM {table}")
    return {r["name"]: r["id"] for r in cur.fetchall()}

def _ensure_lookup(conn, table: str, name: str) -> int:
    conn.execute(f"INSERT OR IGNORE INTO {table}(name) VALUES (?)", (name,))
    cur = conn.execute(f"SELECT id FROM {table} WHERE name=?", (name,))
    return int(cur.fetchone()[0])

def _upsert_network(conn, data: Dict[str, Any]) -> int:
    now = datetime.utcnow().isoformat(timespec="seconds")
    cur = conn.execute("SELECT id FROM networks WHERE frequency=?", (data["frequency"],))
    row = cur.fetchone()
    if row:
        nid = int(row["id"])
        conn.execute("""UPDATE networks SET
            mask=?, unit=?, zone=?, chat_id=?, group_id=?, status_id=?, comment=?, updated_at=?
            WHERE id=?""",
            (data.get("mask"), data["unit"], data["zone"], data["chat_id"], data["group_id"], data["status_id"], data.get("comment"), now, nid)
        )
        return nid
    cur = conn.execute("""INSERT INTO networks
        (frequency, mask, unit, zone, chat_id, group_id, status_id, comment, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?)""",
        (data["frequency"], data.get("mask"), data["unit"], data["zone"], data["chat_id"], data["group_id"], data["status_id"], data.get("comment"), now)
    )
    return int(cur.lastrowid)

def _set_network_tags(conn, network_id: int, tag_names: List[str]):
    conn.execute("DELETE FROM network_tags WHERE network_id=?", (network_id,))
    tag_names = [t.strip() for t in tag_names if t and t.strip()]
    for nm in sorted(set(tag_names)):
        tid = _ensure_lookup(conn, "tags", nm)
        conn.execute("INSERT OR IGNORE INTO network_tags(network_id, tag_id) VALUES (?,?)", (network_id, tid))

def _upsert_etalon(conn, network_id: int, et: Dict[str, Any]):
    now = datetime.utcnow().isoformat(timespec="seconds")
    cur = conn.execute("SELECT id FROM etalons WHERE network_id=?", (network_id,))
    row = cur.fetchone()

    start_date = et.get("start_date")
    start_date_str = start_date.isoformat() if isinstance(start_date, date) else None

    if row:
        eid = int(row["id"])
        conn.execute("""UPDATE etalons SET
            start_date=COALESCE(?, start_date),
            correspondents=COALESCE(?, correspondents),
            callsigns=COALESCE(?, callsigns),
            purpose=COALESCE(?, purpose),
            operation_mode=COALESCE(?, operation_mode),
            traffic_type=COALESCE(?, traffic_type),
            raw_import_text=COALESCE(?, raw_import_text),
            updated_at=?
            WHERE id=?""",
            (start_date_str,
             et.get("correspondents"),
             et.get("callsigns"),
             et.get("purpose"),
             et.get("operation_mode"),
             et.get("traffic_type"),
             et.get("raw_import_text"),
             now, eid)
        )
        return
    conn.execute("""INSERT INTO etalons
        (network_id, start_date, correspondents, callsigns, purpose, operation_mode, traffic_type, raw_import_text, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?)""",
        (network_id, start_date_str, et.get("correspondents"), et.get("callsigns"), et.get("purpose"),
         et.get("operation_mode"), et.get("traffic_type"), et.get("raw_import_text"), now)
    )

def import_freq_table(xlsx_path: str):
    if not os.path.exists(xlsx_path):
        return
    wb = load_workbook(xlsx_path, data_only=True)
    if "freq" not in wb.sheetnames:
        return
    ws = wb["freq"]

    # header row is first row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers.append(str(cell).strip() if cell is not None else "")
    # build index map
    idx = {h: i for i, h in enumerate(headers)}

    def col_by_name(names: List[str]) -> Optional[int]:
        for n in names:
            for h, i in idx.items():
                if h.lower() == n.lower():
                    return i
        return None

    col_freq = col_by_name(["Частота"])
    if col_freq is None:
        return
    col_mask = col_by_name(["Маска_3","Маска"])
    col_unit = col_by_name(["Підрозділ"])
    col_group = col_by_name(["Хто"])
    col_zone = next((i for h,i in idx.items() if h.lower().startswith("зона")), None)
    col_status = col_by_name(["Статус"])
    col_tags = col_by_name(["Теги"])
    col_comment = col_by_name(["Коментар"])
    col_start = col_by_name(["Початок"])

    # detect chat col
    chat_col = None
    for h,i in idx.items():
        if h.lower() in {"чат","джерело","source"}:
            chat_col = i
            break
    if chat_col is None:
        # scan columns for known values in first 200 rows
        for i in range(len(headers)):
            hits = 0
            for r in ws.iter_rows(min_row=2, max_row=201, values_only=True):
                v = r[i] if i < len(r) else None
                if v is None:
                    continue
                if str(v).strip() in CHAT_VALUES:
                    hits += 1
            if hits > 0:
                chat_col = i
                break

    with get_conn() as conn:
        status_map = _get_id_map(conn, "statuses")
        chat_map = _get_id_map(conn, "chats")
        group_map = _get_id_map(conn, "groups")

        for r in ws.iter_rows(min_row=2, values_only=True):
            raw_f = r[col_freq] if col_freq < len(r) else None
            nf = normalize_freq(None if raw_f is None else str(raw_f))
            if not nf:
                continue

            raw_m = r[col_mask] if (col_mask is not None and col_mask < len(r)) else None
            nm = normalize_freq(None if raw_m is None else str(raw_m))

            unit = str(r[col_unit]).strip() if (col_unit is not None and col_unit < len(r) and r[col_unit] is not None) else ""
            zone = str(r[col_zone]).strip() if (col_zone is not None and col_zone < len(r) and r[col_zone] is not None) else ""
            group_name = str(r[col_group]).strip() if (col_group is not None and col_group < len(r) and r[col_group] is not None) else ""
            status_name = str(r[col_status]).strip() if (col_status is not None and col_status < len(r) and r[col_status] is not None) else "Спостерігається"

            if status_name not in status_map:
                status_map[status_name] = _ensure_lookup(conn, "statuses", status_name)

            chat_name = "Очерет"
            if chat_col is not None and chat_col < len(r) and r[chat_col] is not None:
                cn = str(r[chat_col]).strip()
                if cn in CHAT_VALUES:
                    chat_name = cn
            if chat_name not in chat_map:
                chat_map[chat_name] = _ensure_lookup(conn, "chats", chat_name)

            if group_name:
                if group_name not in group_map:
                    group_map[group_name] = _ensure_lookup(conn, "groups", group_name)
            else:
                group_name = next(iter(group_map.keys()))

            comment = str(r[col_comment]).strip() if (col_comment is not None and col_comment < len(r) and r[col_comment] is not None) else None
            if comment == "":
                comment = None

            nid = _upsert_network(conn, {
                "frequency": nf,
                "mask": nm,
                "unit": unit or "",
                "zone": zone or "",
                "chat_id": chat_map[chat_name],
                "group_id": group_map[group_name],
                "status_id": status_map[status_name],
                "comment": comment,
            })

            if col_tags is not None and col_tags < len(r) and r[col_tags] is not None:
                tag_text = str(r[col_tags]).strip()
                if tag_text:
                    parts = [p.strip() for p in re.split(r"[,;]+", tag_text) if p.strip()]
                    _set_network_tags(conn, nid, parts)

            et = {}
            if col_start is not None and col_start < len(r) and r[col_start] is not None:
                sd = _parse_date_any(str(r[col_start]))
                if sd:
                    et["start_date"] = sd
            if et:
                _upsert_etalon(conn, nid, et)

def import_etalon_tabs(xlsx_path: str):
    if not os.path.exists(xlsx_path):
        return
    wb = load_workbook(xlsx_path, data_only=True)
    with get_conn() as conn:
        cur = conn.execute("SELECT id, frequency FROM networks")
        net_by_freq = {row["frequency"]: row["id"] for row in cur.fetchall()}

        for sh in wb.sheetnames:
            if sh in SKIP_SHEETS:
                continue
            nf = _sheet_name_is_freq(sh)
            if not nf:
                continue
            nid = net_by_freq.get(nf)
            if not nid:
                continue
            data = _parse_etalon_sheet(wb, sh)
            _upsert_etalon(conn, nid, data)

def main():
    init_db()
    import_freq_table(settings.freq_xlsx)
    import_etalon_tabs(settings.freq_xlsx)
    if settings.etalon_xlsx:
        import_etalon_tabs(settings.etalon_xlsx)

if __name__ == "__main__":
    main()
